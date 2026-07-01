"""The multi-skill Excel export builds a valid workbook whose numbers match the engine."""
import io

import openpyxl

from modules.calculations.engine import compute_multi_skill_model
from modules.outputs.multi_excel_export import generate_multi_excel_report


def _skill(sid, name, genus, levels, arch):
    return {"id": sid, "name": name, "genus_category": genus, "active_levels": levels,
            "has_architect": arch > 0, "architect_pct": float(arch), "coverage_model": "24×7",
            "visible": True, "level_visible": {},
            "role_buffers": {"L1": 20.0, "L2": 20.0, "L3": 20.0, "Architect": 0.0},
            "workload": {"service_requests": {"All": {"count": 20, "minutes": 45,
                                                      "L1_pct": 0, "L2_pct": 60, "L3_pct": 40}}},
            "patching": None, "activities": []}


def _state():
    return {
        "skills": [_skill("s1", "Cloud Operations", "CloudOps", ["L2", "L3"], 25),
                   _skill("s2", "DevOps", "CloudOps", ["L2", "L3"], 25)],
        "resource_sharing": [], "sdm_overhead_pct": 5.0, "sdm_rate_inr": 2200,
        "rates_by_category": {"InfraOps": {"L1": 750, "L2": 1100, "L3": 1600, "Architect": 2400},
                              "CloudOps": {"L1": 1000, "L2": 1500, "L3": 2200, "Architect": 3000}},
        "contingency_pct": 10.0, "monthly_working_hours": 160.0, "productive_utilisation": 75.0,
        "fte_basis": "rounded", "target_margin_pct": 25.0, "context_switch_pct": 10.0,
        "enforce_min_shift": False, "custom_hours_per_day": 8, "custom_days_per_week": 5,
        "delivery_country": "India", "delivery_location": "Noida",
    }


def test_workbook_has_expected_sheets():
    wb = openpyxl.load_workbook(io.BytesIO(generate_multi_excel_report(_state())))
    for name in ["Executive Summary", "Skills", "Effort Build-up", "Team (FTE)",
                 "Rates", "Optimization", "Workload Detail", "Inputs"]:
        assert name in wb.sheetnames


def test_exec_total_fte_matches_engine():
    st = _state()
    model = compute_multi_skill_model(st)
    wb = openpyxl.load_workbook(io.BytesIO(generate_multi_excel_report(st)))
    ex = wb["Executive Summary"]
    labelled = {ex.cell(r, 1).value: ex.cell(r, 2).value for r in range(6, 20)}
    assert labelled["TOTAL FTE"] == round(model["total_fte"], 1)
    assert labelled["MONTHLY SELLING PRICE"] == model["price_result"]["selling_price"]


def test_skills_cost_total_reconciles():
    st = _state()
    model = compute_multi_skill_model(st)
    wb = openpyxl.load_workbook(io.BytesIO(generate_multi_excel_report(st)))
    ws = wb["Skills"]
    # find the TOTAL row's Monthly Cost (last column, col 9)
    total_cost = None
    for r in range(6, ws.max_row + 1):
        if ws.cell(r, 1).value == "TOTAL":
            total_cost = ws.cell(r, 9).value
    per_skill_cost = round(sum(ps.get("cost", 0.0) for ps in model["per_skill"].values()))
    assert total_cost == per_skill_cost
