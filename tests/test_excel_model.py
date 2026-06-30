"""Regression guard for the Editable Excel workbook: generate it, recalculate every
formula with the `formulas` engine, and assert key cells match compute_full_model.

This is what catches a class of bug the static review missed (e.g. an unqualified
cross-sheet SUMPRODUCT that silently self-references the wrong sheet). Skips cleanly
if the optional `formulas` dependency isn't installed."""
import io
import pandas as pd
import pytest

formulas = pytest.importorskip("formulas")

from modules.calculations.engine import compute_full_model, resolve_role_rates


class _SS(dict):
    def __getattr__(self, k):
        try: return self[k]
        except KeyError: raise AttributeError(k)
    def __setattr__(self, k, v): self[k] = v


def _cat(rows):
    return {lbl: dict(count=c, minutes=m, L1_pct=l1, L2_pct=l2, L3_pct=l3,
                      L1_buffer=20, L2_buffer=20, L3_buffer=20)
            for lbl, c, m, l1, l2, l3 in rows}


def _state():
    rc = pd.DataFrame([
        {"country": "India", "location": "Pune", "genus": g, "hourly rate": r, "rate currency": "INR"}
        for g, r in [("2.1-INFRAOPS", 700), ("2.3-INFRAOPS", 1100), ("3.2-INFRAOPS", 1600),
                     ("4.1-INFRAOPS", 2400), ("4.1-DELIVERY-ITIL", 2200), ("4.2-DELIVERY-ITIL", 3000)]])
    s = {
        "alerts": _cat([("Critical", 100, 30, 60, 30, 10), ("High", 200, 20, 80, 20, 0)]),
        "service_requests": _cat([("Low", 300, 15, 90, 10, 0)]),
        "incidents": _cat([("High", 20, 120, 20, 50, 30)]),
        "changes": _cat([("Normal", 40, 60, 30, 50, 20)]),
        "patching_included": "Yes", "num_servers": 20, "patching_method": "Manual",
        "manual_effort_per_server": 45, "auto_effort_per_server": 30, "patch_error_rate": 10,
        "patching_role": "L2",
        # an additional activity with a NON-zero L2 split — this is what surfaced the
        # cross-sheet SUMPRODUCT bug; keep it in the regression scenario.
        "additional_activities": [{"name": "RCA", "hours": 12.0, "custom": True, "auto": False,
                                   "dist": {"L1": 0, "L2": 30, "L3": 50, "Architect": 20, "SDM": 0}}],
        "contingency_pct": 10.0, "overhead_pcts": {"Architect": 5.0, "SDM": 5.0},
        "coverage_model": "24×7", "monthly_working_hours": 160.0, "productive_utilisation": 75.0,
        "delivery_country": "India", "delivery_location": "Pune",
        "role_genus": {"L1": "2.1-INFRAOPS", "L2": "2.3-INFRAOPS", "L3": "3.2-INFRAOPS",
                       "Architect": "4.1-INFRAOPS", "SDM": "4.1-DELIVERY-ITIL"},
        "additional_costs": [{"name": "Tools", "cost": 20000.0, "custom": True}],
        "sla_provision_included": "Yes", "sla_provision_pct": 2.0,
        "target_margin_pct": 20.0, "reporting_currency": "INR", "fte_basis": "rounded",
        "exchange_rates": {}, "transition_total_cost": 0.0, "rate_card_df": rc,
        "transition_planner": {
            "enabled": True, "total_weeks": 4,
            "phases": [{"id": "p1", "name": "Assess", "weeks": 2}, {"id": "p2", "name": "KT", "weeks": 2}],
            "resources": [{"id": "a", "role": "L2", "count": 2}, {"id": "b", "role": "L3", "count": 1}],
            "allocation": {"a": {"1": 1.0, "2": 1.0, "3": 0.5, "4": 0.5},
                           "b": {"1": 0.5, "2": 0.5, "3": 1.0, "4": 1.0}},
            "treatment": "recurring", "amortisation_months": 12},
    }
    s["role_rates_inr"] = resolve_role_rates(rc, s["role_genus"], "India", "Pune", {})
    return s


@pytest.fixture
def recalced(tmp_path, monkeypatch):
    import streamlit as st
    state = _state()
    model = compute_full_model(state)
    ss = _SS(state); ss["_model"] = model
    ss["workload_totals"] = {"alerts": 300, "incidents": 20, "service_requests": 300, "changes": 40}
    monkeypatch.setattr(st, "session_state", ss, raising=False)
    from modules.outputs.excel_model import generate_excel_model
    data = generate_excel_model()
    path = (tmp_path / "m.xlsx"); path.write_bytes(data)
    xl = formulas.ExcelModel().loads(str(path)).finish()
    sol = xl.calculate()
    fname = path.name

    def val(sheet, coord):
        v = sol[f"'[{fname}]{sheet.upper()}'!{coord}"].value
        try: return float(v[0, 0])
        except Exception: return float(v)
    return model, val, io.BytesIO(data)


def _find_row(wb_sheet, label, col=1, value_col=2):
    for row in wb_sheet.iter_rows():
        if isinstance(row[col - 1].value, str) and row[col - 1].value == label:
            return row[value_col - 1].coordinate
    raise AssertionError(f"label {label!r} not found")


def test_workbook_recalculates_to_engine(recalced):
    import openpyxl
    model, val, data = recalced
    wb = openpyxl.load_workbook(data)

    # Role hours (Effort sheet) — guards the activity cross-sheet SUMPRODUCT bug
    eff = wb["5 Effort"]
    for i, role in enumerate(["L1", "L2", "L3", "Architect", "SDM"]):
        coord = _find_row(eff, role)
        assert val("5 Effort", coord) == pytest.approx(model["role_hours"][role], rel=0.01), role

    # Transition + price headline
    tr = wb["Transition"]
    assert val("Transition", _find_row(tr, "Total transition cost")) == pytest.approx(model["transition"]["total"], rel=0.01)
    assert val("Transition", _find_row(tr, "Monthly recurring (÷ months)")) == pytest.approx(
        model["transition"]["monthly_recurring"], rel=0.01)

    co = wb["8 Costing"]
    assert val("8 Costing", _find_row(co, "MONTHLY PRICE incl. transition")) == pytest.approx(
        model["monthly_price_with_transition"], rel=0.01)
    assert val("8 Costing", _find_row(co, "Monthly selling price")) == pytest.approx(
        model["price_result"]["selling_price"], rel=0.01)


def test_workbook_has_all_sheets(recalced):
    import openpyxl
    _, _, data = recalced
    wb = openpyxl.load_workbook(data)
    for s in ["Summary", "Inputs", "Transition", "8 Costing", "Dashboard"]:
        assert s in wb.sheetnames
