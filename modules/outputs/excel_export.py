"""Excel export — multi-sheet workbook. Reads from the unified compute model."""
import io
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
from config.settings import (
    ALL_ROLES, APP_NAME, DEFAULT_ROLE_BUFFER_PCT, REPORTING_CURRENCIES, hx,
)
from modules.state.session_manager import run_model

# Brand tokens (shared with the web app and PDF via config.settings.THEME)
NAVY = hx("navy"); BLUE = hx("badge"); LB = hx("tint"); ACCENT = hx("primary")
LGRAY = "F2F2F2"; WHITE = "FFFFFF"

# ── Number formats ─────────────────────────────────────────────────────────────
F_INR  = "#,##0"        # currency / counts (INR column)
F_CUR2 = "#,##0.00"     # reporting-currency value
F_NUM1 = "#,##0.0"      # hours
F_FTE  = "0.0"          # FTE
F_RAW  = "0.000"        # raw FTE
F_PCT  = '0.0"%"'       # percentage stored as e.g. 20.0 → "20.0%"
F_PCT0 = '0"%"'         # whole-number percentage → "80%"


def _unit_fmt(u):
    """Map an Exec/Audit 'unit' label to the right Excel number format."""
    u = (u or "").strip()
    if u in ("Hrs/Month", "Hrs"): return F_NUM1
    if u == "FTE": return F_FTE
    if u == "%":   return F_PCT
    if u == "INR": return F_INR
    if u in REPORTING_CURRENCIES: return F_CUR2
    return None

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(c="000000", bold=False, sz=10): return Font(name="Calibri", color=c, bold=bold, size=sz)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def _hrow(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = _fill(NAVY); c.font = _font(WHITE, True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        if widths and i <= len(widths): _cw(ws, i, widths[i-1])

def _drow(ws, row, vals, total=False, fmts=None):
    bg = _fill(LB) if total else None
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = _font(bold=total)
        c.border = _border()
        c.alignment = Alignment(horizontal="right" if isinstance(v, (int, float)) else "left")
        if fmts and i <= len(fmts) and fmts[i-1] and isinstance(v, (int, float)):
            c.number_format = fmts[i-1]
        if bg: c.fill = bg

def _title(ws, t, sub=""):
    ws.cell(1,1,t).font = Font("Calibri", bold=True, size=13, color=NAVY)
    if sub: ws.cell(2,1,sub).font = Font("Calibri", italic=True, size=10, color="767676")
    ws.cell(3,1,"Generated:").font = Font("Calibri", size=9, color="767676")
    dc = ws.cell(3,2, date.today())          # real Date cell, not a string
    dc.number_format = "yyyy-mm-dd"
    dc.font = Font("Calibri", size=9, color="767676")

def _build_exec(wb, model):
    ws = wb.create_sheet("Executive Summary")
    _title(ws, f"{APP_NAME} — Executive Summary")
    cost = model["cost_result"]; price = model["price_result"]
    te = model["total_effort"]; base = model["base_effort"]
    _hrow(ws, 5, ["Metric", "Value", "Unit"], [35, 22, 15])
    rows = [
        ("Base Operational Effort",   round(base, 1),  "Hrs/Month"),
        ("Contingency Hours",         round(te-base,1),"Hrs/Month"),
        ("Total Operational Effort",  round(te, 1),    "Hrs/Month"),
        ("Total FTE Required",        round(model["total_fte"],1),"FTE"),
        ("Coverage Model",            st.session_state.get("coverage_model","—"),""),
        ("Total Resource Cost (INR)", cost.get("resource_cost",0), "INR"),
        ("Additional Expenses (INR)", cost.get("additional_expenses",0),"INR"),
        ("Total Delivery Cost (INR)", cost.get("total_delivery_cost",0),"INR"),
        ("Gross Margin %",            price.get("margin_pct",0), "%"),
        ("Monthly Selling Price (INR)",price.get("selling_price",0),"INR"),
        ("One-Time Transition Cost (INR)", model["transition_cost"], "INR"),
        ("Reporting Currency",        model["reporting_currency"], ""),
        (f"Monthly Selling Price ({model['reporting_currency']})",
         round(model["selling_price_converted"], 2), model["reporting_currency"]),
    ]
    for i,(m,v,u) in enumerate(rows,6):
        _drow(ws,i,[m,v,u], total=(str(m).isupper()), fmts=[None, _unit_fmt(u), None])

def _build_effort(wb, model):
    ws = wb.create_sheet("Effort Breakdown")
    _title(ws, "Effort Breakdown")
    rh = model["role_hours"]; te = model["total_effort"]
    _hrow(ws,5,["Role","Effort Hours","% of Total Effort"],[18,16,20])
    EFFORT_FMTS = [None, F_NUM1, F_PCT]
    for i,role in enumerate(ALL_ROLES,6):
        h = rh.get(role,0); pct = h/te*100 if te>0 else 0
        _drow(ws,i,[role,round(h,1),round(pct,1)], fmts=EFFORT_FMTS)
    _drow(ws,6+len(ALL_ROLES),["TOTAL",round(te,1),100.0],total=True, fmts=EFFORT_FMTS)

def _build_resolution(wb, model):
    ws = wb.create_sheet("Resolution Detail")
    _title(ws, "Resolution Detail — Tickets, Minutes & Role Hours")
    _hrow(ws,5,[
        "Category","Severity/Type","Count","Min/Ticket","Total Hrs",
        "L1 %","L1 Buf%","L1 Hrs","L2 %","L2 Buf%","L2 Hrs","L3 %","L3 Buf%","L3 Hrs"
    ],[18,14,10,12,12,8,9,10,8,9,10,8,9,10])
    from config.settings import CATEGORY_SUBLABELS
    CATS = [
        ("alerts",           "Monitoring Alerts"),
        ("service_requests", "Service Requests"),
        ("incidents",        "Incidents"),
        ("changes",          "Change Requests"),
    ]
    row_num = 6
    for cat_key, cat_label in CATS:
        cat_data = st.session_state.get(cat_key, {})
        for label in CATEGORY_SUBLABELS[cat_key]:
            row = cat_data.get(label, {})
            cnt  = row.get("count", 0); mins = row.get("minutes", 0)
            l1p  = row.get("L1_pct", 0); l2p = row.get("L2_pct", 0); l3p = row.get("L3_pct", 0)
            l1b = row.get("L1_buffer", DEFAULT_ROLE_BUFFER_PCT)
            l2b = row.get("L2_buffer", DEFAULT_ROLE_BUFFER_PCT)
            l3b = row.get("L3_buffer", DEFAULT_ROLE_BUFFER_PCT)
            total_h = (cnt * mins) / 60.0
            _drow(ws, row_num, [
                cat_label, label, cnt, round(mins,0), round(total_h,2),
                l1p, l1b, round(total_h*l1p/100*(1+l1b/100),2),
                l2p, l2b, round(total_h*l2p/100*(1+l2b/100),2),
                l3p, l3b, round(total_h*l3p/100*(1+l3b/100),2),
            ], fmts=[None, None, F_INR, F_INR, F_NUM1,
                     F_PCT0, F_PCT0, F_NUM1, F_PCT0, F_PCT0, F_NUM1, F_PCT0, F_PCT0, F_NUM1])
            row_num += 1

def _build_fte(wb, model):
    ws = wb.create_sheet("FTE Summary")
    _title(ws, "FTE Summary")
    fte = model["fte_result"]
    _hrow(ws,5,["Role","Effort Hours","Raw FTE","Cov. Applied","Final FTE"],[14,14,12,16,12])
    FTE_FMTS = [None, F_NUM1, F_RAW, None, F_FTE]
    for i,role in enumerate(ALL_ROLES,6):
        r = fte.get(role,{})
        _drow(ws,i,[role,round(r.get("hours",0),1),round(r.get("raw_fte",0),3),"Yes" if r.get("coverage_applied") else "No",round(r.get("final_fte",0),1)], fmts=FTE_FMTS)
    _drow(ws,6+len(ALL_ROLES),["TOTAL","","","",round(model["total_fte"],1)],total=True, fmts=FTE_FMTS)

def _build_costs(wb, model):
    ws = wb.create_sheet("Resource Costs")
    _title(ws,"Resource Cost Summary")
    rc = model["resource_costs"]
    _hrow(ws,5,["Role","Genus","Required FTE","Billed Hours","Rate (INR/hr)","Cost (INR)"],[12,22,14,14,16,18])
    COST_FMTS = [None, None, F_FTE, F_INR, F_INR, F_INR]
    for i,role in enumerate(ALL_ROLES,6):
        r = rc.get(role,{})
        _drow(ws,i,[role,r.get("genus","—"),round(r.get("fte",0),1),round(r.get("billed_hours",0),0),round(r.get("rate_inr",0),0),round(r.get("cost_inr",0),0)], fmts=COST_FMTS)
    _drow(ws,6+len(ALL_ROLES),["TOTAL","","","","",round(model["total_resource_cost"],0)],total=True, fmts=COST_FMTS)

def _build_financial(wb, model):
    ws = wb.create_sheet("Financial Model")
    _title(ws,"Financial Model")
    cost = model["cost_result"]; price = model["price_result"]
    _hrow(ws,5,["Component","Amount (INR)"],[38,20])
    rows = [
        ("Total Resource Cost",       cost.get("resource_cost",0),       False),
        ("Total Additional Expenses", cost.get("additional_expenses",0), False),
        ("SLA Penalty Provision",     cost.get("sla_provision",0),       False),
        ("TOTAL DELIVERY COST",       cost.get("total_delivery_cost",0), True),
        ("Gross Margin %",            price.get("margin_pct",0),         False),
        ("Gross Profit",              price.get("gross_profit",0),       False),
        ("MONTHLY SELLING PRICE",     price.get("selling_price",0),      True),
        ("ONE-TIME TRANSITION COST",  model["transition_cost"],          True),
    ]
    for i,(lbl,val,tot) in enumerate(rows,6):
        is_pct = "%" in lbl
        v = round(val, 1) if is_pct else round(val, 0)
        _drow(ws,i,[lbl,v],total=tot, fmts=[None, F_PCT if is_pct else F_INR])

def _build_audit(wb, model):
    ws = wb.create_sheet("Inputs Audit")
    _title(ws,"All User Inputs — Audit Trail")
    _hrow(ws,5,["Step","Input Field","Value","Unit"],[8,40,22,15])
    audit=[
        (3,"Patching Included",st.session_state.get("patching_included",""),""),
        (3,"Number of Servers",st.session_state.get("num_servers",0),""),
        (3,"Patching Method",st.session_state.get("patching_method",""),""),
        (5,"Contingency %",st.session_state.get("contingency_pct",0),"%"),
        (6,"Coverage Model",st.session_state.get("coverage_model",""),""),
        (6,"Monthly Working Hours",st.session_state.get("monthly_working_hours",0),"Hrs"),
        (6,"Productive Utilisation",st.session_state.get("productive_utilisation",0),"%"),
        (1,"Delivery Country",st.session_state.get("delivery_country","India"),""),
        (1,"Delivery Location",st.session_state.get("delivery_location","—") or "—",""),
        (8,"Reporting Currency",model["reporting_currency"],""),
        (8,"Target Gross Margin",st.session_state.get("target_margin_pct",0),"%"),
        (8,"Transition Included",st.session_state.get("transition_included",""),""),
        (8,"Transition Total Cost",st.session_state.get("transition_total_cost",0),"INR"),
        (8,"SLA Provision Included",st.session_state.get("sla_provision_included",""),""),
        (8,"SLA Provision %",st.session_state.get("sla_provision_pct",0),"%"),
    ]
    for role, pct in (st.session_state.get("overhead_pcts",{}) or {}).items():
        audit.append((5,f"Overhead % — {role}",pct,"%"))
    for i,row in enumerate(audit,6):
        step, field, value, unit = row
        _drow(ws,i,list(row), fmts=[None, None, _unit_fmt(unit), None])

def generate_excel_report() -> bytes:
    model = st.session_state.get("_model") or run_model()
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _build_exec(wb, model); _build_effort(wb, model); _build_resolution(wb, model)
    _build_fte(wb, model); _build_costs(wb, model); _build_financial(wb, model); _build_audit(wb, model)
    tab_colors = [NAVY, hx("teal_dark"), BLUE, ACCENT, hx("success"), hx("text_muted")]
    for i,ws in enumerate(wb.worksheets):
        ws.sheet_properties.tabColor = tab_colors[i % len(tab_colors)]
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
