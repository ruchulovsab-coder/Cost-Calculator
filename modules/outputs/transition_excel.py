"""Excel export for the Transition Strategy — themed, proposal-ready appendix.

Sheets: Timeline (phase Gantt-style bands + milestones), Phase Activities, Skill-wise Plan,
Acceptance & Sign-off (per-skill exit/sign-off gates + fillable open-items/risk register + named
sign-off block), RACI, Deliverables, RAID Register (seeded R/A/D + Issue rows), Governance & Comms.
Presentation values (not a recalc model), app theme, no logo (parity with the other formula-driven
exports)."""
from __future__ import annotations

import io
from datetime import date
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import hx

NAVY = hx("navy"); TEAL = hx("teal_dark"); TINT = hx("tint"); MUTED = hx("text_muted")
ACCENT = hx("accent_light"); AMBER = "FBEED9"
_thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
RACI_FILL = {"R": "D6F0ED", "A": ACCENT, "C": "EAF3F4", "I": "F4F6F7"}


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _title(ws, r, text, size=13):
    ws.cell(r, 1, text).font = Font(bold=True, color=NAVY, size=size)


def _hdr(ws, r, headers, c0=1):
    for j, h in enumerate(headers, start=c0):
        c = ws.cell(r, j, h); c.fill = _fill(NAVY); c.border = BORDER
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cell(ws, r, j, val, bold=False, wrap=False, fill=None, center=False):
    c = ws.cell(r, j, val); c.border = BORDER
    c.font = Font(size=10, bold=bold)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="top", wrap_text=wrap)
    if fill:
        c.fill = _fill(fill)


def _bullets(items: List[str]) -> str:
    return "\n".join("• " + str(i) for i in (items or [])) or "—"


def build_transition_workbook(plan: Dict[str, Any], project: str = "") -> bytes:
    wb = openpyxl.Workbook()

    # ── Timeline ──
    ws = wb.active; ws.title = "Timeline"; ws.sheet_view.showGridLines = False
    r = 1
    _title(ws, r, f"Transition Strategy{(' — ' + project) if project else ''}"); r += 2
    ws.cell(r, 1, f"Start {plan['start'].isoformat()} · span {plan['span_weeks']:g} weeks · "
                  f"customer TZ {plan.get('customer_tz','')}").font = Font(size=10, color=MUTED); r += 1
    ws.cell(r, 1, "Foundation: " + plan.get("foundation", "")).font = Font(size=10, italic=True,
                                                                           color=MUTED); r += 2
    _hdr(ws, r, ["Phase", "ITIL Band", "Start", "End", "Weeks", "Milestone / Gate"]); r += 1
    ms_by_phase = {m["phase"]: m for m in plan.get("milestones", [])}
    for row in plan.get("timeline", []):
        ms = ms_by_phase.get(row["name"])
        _cell(ws, r, 1, row["name"], bold=True)
        _cell(ws, r, 2, row["band"])
        _cell(ws, r, 3, row["start"].isoformat(), center=True)
        _cell(ws, r, 4, row["end"].isoformat(), center=True)
        _cell(ws, r, 5, f"{row['duration_weeks']:g}", center=True)
        _cell(ws, r, 6, (f"{ms['id']}: {ms['gate']}" if ms else ""), wrap=True,
              fill=AMBER if ms else None)
        r += 1
    for j, w in enumerate([34, 18, 12, 12, 7, 46], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ── Phase Activities ──
    ws = wb.create_sheet("Phase Activities"); ws.sheet_view.showGridLines = False
    r = 1; _title(ws, r, "Phase Activities"); r += 1
    _hdr(ws, r, ["Phase", "Objectives", "Deliverables", "Entry", "Exit", "Risks",
                 "Dependencies", "Customer", "Nagarro"]); r += 1
    for p in plan.get("phase_activities", []):
        _cell(ws, r, 1, p["name"], bold=True, wrap=True)
        for j, k in enumerate(["objectives", "deliverables", "entry", "exit", "risks",
                               "dependencies", "customer_resp", "nagarro_resp"], start=2):
            _cell(ws, r, j, _bullets(p.get(k, [])), wrap=True)
        r += 1
    ws.column_dimensions["A"].width = 22
    for j in range(2, 10):
        ws.column_dimensions[get_column_letter(j)].width = 30

    # ── Skill-wise Plan ──
    ws = wb.create_sheet("Skill-wise Plan"); ws.sheet_view.showGridLines = False
    r = 1; _title(ws, r, "Skill-wise Transition Plan"); r += 1
    _hdr(ws, r, ["Skill", "Family", "Levels", "Knowledge Transition", "Shadow", "Reverse Shadow",
                 "Stabilization", "Exit Criteria", "Sign-off Criteria"]); r += 1
    for sp in plan.get("skill_plans", []):
        _cell(ws, r, 1, sp["skill"], bold=True, wrap=True)
        _cell(ws, r, 2, sp.get("family_label", "General"), wrap=True)
        _cell(ws, r, 3, ", ".join(sp["levels"]) or "—", wrap=True)
        for j, k in enumerate(["knowledge_transition", "shadow", "reverse_shadow", "stabilization",
                               "exit_criteria", "signoff_criteria"], start=4):
            _cell(ws, r, j, _bullets(sp.get(k, [])), wrap=True)
        r += 1
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    for j in range(4, 10):
        ws.column_dimensions[get_column_letter(j)].width = 32

    # ── Acceptance & Sign-off (per skill) ──
    ws = wb.create_sheet("Acceptance & Sign-off"); ws.sheet_view.showGridLines = False
    r = 1
    _title(ws, r, "Acceptance & Sign-off — Exit / Go-Live Gates per Skill"); r += 1
    ws.cell(r, 1, "Templates — completed during the transition. Every open item must carry a named owner "
                  "and target date and be agreed by both parties; residual risk must be accepted by both "
                  "parties before sign-off.").font = Font(size=9, italic=True, color=MUTED); r += 2
    oi_cols = plan.get("open_items_columns", []) or ["#"]
    ncol = len(oi_cols)
    for sp in plan.get("skill_plans", []):
        _title(ws, r, f"{sp['skill']}  —  {sp.get('family_label', 'General')}", 11); r += 1
        for label, key, fill in [("Exit criteria (KT/Shadow gate)", "exit_criteria", TINT),
                                 ("Sign-off criteria (Go-Live gate)", "signoff_criteria", TINT)]:
            _cell(ws, r, 1, label, bold=True, fill=fill, wrap=True)
            _cell(ws, r, 2, _bullets(sp.get(key, [])), wrap=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncol); r += 1
        _cell(ws, r, 1, "Critical readiness check", bold=True, fill=ACCENT, wrap=True)
        _cell(ws, r, 2, sp.get("family_critical_check", ""), wrap=True, fill=ACCENT)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncol); r += 2
        # Fillable open-items / residual-risk register
        _cell(ws, r, 1, "Open Items & Residual Risk register", bold=True); r += 1
        _hdr(ws, r, oi_cols); r += 1
        for i in range(6):
            _cell(ws, r, 1, i + 1, center=True)
            for j in range(2, ncol + 1):
                _cell(ws, r, j, "")
            r += 1
        r += 1
        # Named sign-off block
        _cell(ws, r, 1, "Sign-off — recorded at the gate", bold=True); r += 1
        _hdr(ws, r, ["Party", "Role", "Name", "Signature", "Date"]); r += 1
        for party, role in plan.get("signoff_signatories", []):
            _cell(ws, r, 1, party); _cell(ws, r, 2, role, wrap=True)
            _cell(ws, r, 3, ""); _cell(ws, r, 4, ""); _cell(ws, r, 5, ""); r += 1
        _cell(ws, r, 1, plan.get("signoff_decision", ""), bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol); r += 3
    ws.column_dimensions["A"].width = 24
    for j in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(j)].width = 18

    # ── RACI ──
    ws = wb.create_sheet("RACI"); ws.sheet_view.showGridLines = False
    r = 1; _title(ws, r, "RACI Matrix (R = Responsible · A = Accountable · C = Consulted · I = Informed)")
    r += 1
    roles = plan.get("roles_customer", []) + plan.get("roles_nagarro", [])
    _hdr(ws, r, ["Activity"] + roles); r += 1
    for row in plan.get("raci", []):
        _cell(ws, r, 1, row["activity"], wrap=True)
        for j, role in enumerate(roles, start=2):
            v = row["raci"].get(role, "")
            _cell(ws, r, j, v, center=True, bold=(v == "A"), fill=RACI_FILL.get(v))
        r += 1
    ws.column_dimensions["A"].width = 40
    for j in range(2, len(roles) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 10

    # ── Deliverables ──
    ws = wb.create_sheet("Deliverables"); ws.sheet_view.showGridLines = False
    r = 1; _title(ws, r, "Deliverables, Gates & Best-practice Artifacts"); r += 1
    _hdr(ws, r, ["Phase", "Key Deliverables", "Exit / Quality Gate", "Milestone"]); r += 1
    for d in plan.get("deliverables", []):
        _cell(ws, r, 1, d["phase"], bold=True, wrap=True)
        _cell(ws, r, 2, _bullets(d.get("deliverables", [])), wrap=True)
        _cell(ws, r, 3, _bullets(d.get("exit", [])), wrap=True)
        _cell(ws, r, 4, d.get("milestone") or "", center=True, fill=AMBER if d.get("milestone") else None)
        r += 1
    r += 1
    _title(ws, r, "Best-practice artifacts", 11); r += 1
    for a in plan.get("best_practice_artifacts", []):
        ws.cell(r, 1, "• " + a).font = Font(size=10); r += 1
    for j, w in enumerate([22, 44, 40, 12], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ── RAID Register ──
    ws = wb.create_sheet("RAID Register"); ws.sheet_view.showGridLines = False
    r = 1; _title(ws, r, "RAID Register — Risks · Assumptions · Issues · Dependencies"); r += 1
    ws.cell(r, 1, "Risks/Dependencies seeded from the phase plan; Assumptions listed; Issues logged "
                  "during execution. Complete Owner / Likelihood-Impact / Response / Status during "
                  "the transition.").font = Font(size=9, italic=True, color=MUTED); r += 2
    raid_cols = plan.get("raid_columns", [])
    _hdr(ws, r, raid_cols); r += 1
    raid_fill = {"Risk": "FBEED9", "Dependency": "EAF3F4", "Assumption": "EDF3E6", "Issue": "FDE7E7"}
    for i, item in enumerate(plan.get("raid_register", []), start=1):
        _cell(ws, r, 1, i, center=True)
        _cell(ws, r, 2, item["type"], fill=raid_fill.get(item["type"]))
        _cell(ws, r, 3, item["description"], wrap=True)
        _cell(ws, r, 4, item["phase"], wrap=True)
        for j in range(5, len(raid_cols) + 1):
            _cell(ws, r, j, "")
        r += 1
    for i in range(3):                       # blank rows for Issues logged during execution
        _cell(ws, r, 1, "")
        for j in range(2, len(raid_cols) + 1):
            _cell(ws, r, j, "")
        r += 1
    for j, w in enumerate([4, 14, 52, 20, 16, 18, 34, 12], start=1):
        if j <= len(raid_cols):
            ws.column_dimensions[get_column_letter(j)].width = w

    # ── Governance & Comms ──
    ws = wb.create_sheet("Governance & Comms"); ws.sheet_view.showGridLines = False
    r = 1; _title(ws, r, "Governance & Communication Cadence"); r += 1
    gov_cols = plan.get("governance_columns", [])
    _hdr(ws, r, gov_cols); r += 1
    for g in plan.get("governance_cadence", []):
        _cell(ws, r, 1, g["forum"], bold=True, wrap=True)
        _cell(ws, r, 2, g["cadence"], center=True)
        _cell(ws, r, 3, g["participants"], wrap=True)
        _cell(ws, r, 4, g["purpose"], wrap=True)
        r += 1
    for j, w in enumerate([30, 20, 40, 46], start=1):
        if j <= len(gov_cols):
            ws.column_dimensions[get_column_letter(j)].width = w

    # ── Advisories (if any) appended to Timeline sheet ──
    if plan.get("advisories"):
        ws = wb["Timeline"]; rr = ws.max_row + 2
        _title(ws, rr, "Advisories (informational — do not affect commercials)", 11); rr += 1
        for a in plan["advisories"]:
            c = ws.cell(rr, 1, "⚠  " + a); c.fill = _fill(AMBER)
            c.font = Font(size=10); c.alignment = Alignment(wrap_text=True, vertical="center")
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6); rr += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_transition_cost_workbook(res: Dict[str, Any], project: str = "") -> bytes:
    """Excel appendix for the multi-skill Transition Cost — effort & cost by skill, by level,
    SDM, and totals. Presentation values (already computed by transition.costing)."""
    LV = ("L1", "L2", "L3", "Architect")
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Transition Cost"; ws.sheet_view.showGridLines = False
    r = 1
    _title(ws, r, f"Transition Cost{(' — ' + project) if project else ''}"); r += 1
    ws.cell(r, 1, f"Duration {res.get('weeks', 0):g} weeks (to Go-Live) · effective "
                  f"{res.get('effective_weeks', 0):g} weeks · {res.get('weekly_hours', 0):g} hrs/week/seat · "
                  "separate one-time line (does not affect the monthly run-rate)."
            ).font = Font(size=10, italic=True, color=MUTED); r += 2

    _hdr(ws, r, ["Skill", "Family", "Transition Team", "FTE", "Hours", "Cost"]); r += 1
    for sp in res.get("per_skill", {}).values():
        team = " · ".join(f"{lvl} {sp['levels'][lvl]['seats']}" for lvl in LV if lvl in sp["levels"]) or "—"
        _cell(ws, r, 1, sp["name"], bold=True)
        _cell(ws, r, 2, sp.get("genus_category", ""))
        _cell(ws, r, 3, team, wrap=True)
        _cell(ws, r, 4, round(sp["fte"], 2), center=True)
        _cell(ws, r, 5, round(sp["hours"]), center=True)
        _cell(ws, r, 6, round(sp["cost"]), center=True)
        r += 1
    sdm = res.get("sdm", {})
    _cell(ws, r, 1, "SDM (engagement)", bold=True); _cell(ws, r, 2, "—")
    _cell(ws, r, 3, f"{sdm.get('fte', 0):.2f} FTE")
    _cell(ws, r, 4, round(sdm.get("fte", 0), 2), center=True)
    _cell(ws, r, 5, round(sdm.get("hours", 0)), center=True)
    _cell(ws, r, 6, round(sdm.get("cost", 0)), center=True); r += 1
    _cell(ws, r, 1, "Total", bold=True, fill=ACCENT); _cell(ws, r, 2, "", fill=ACCENT)
    _cell(ws, r, 3, "", fill=ACCENT)
    _cell(ws, r, 4, round(res.get("total_fte", 0), 2), center=True, bold=True, fill=ACCENT)
    _cell(ws, r, 5, round(res.get("total_hours", 0)), center=True, bold=True, fill=ACCENT)
    _cell(ws, r, 6, round(res.get("total_cost", 0)), center=True, bold=True, fill=ACCENT); r += 2

    _title(ws, r, "By level", 11); r += 1
    _hdr(ws, r, ["Level", "Seats", "Hours", "Cost"]); r += 1
    for lvl in LV:
        d = res.get("by_level", {}).get(lvl, {})
        if not d or d.get("seats", 0) <= 0:
            continue
        _cell(ws, r, 1, lvl, bold=True)
        _cell(ws, r, 2, d["seats"], center=True)
        _cell(ws, r, 3, round(d["hours"]), center=True)
        _cell(ws, r, 4, round(d["cost"]), center=True); r += 1

    for j, w in enumerate([26, 16, 28, 10, 12, 16], start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
