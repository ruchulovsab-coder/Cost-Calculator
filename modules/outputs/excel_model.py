"""
Editable Excel model — a fully formula-driven replica of the app.

ALL user inputs live on ONE sheet ("Inputs") as the only editable (yellow) cells —
every box, radio and checkbox from the application: workload volumetrics, patching,
activities, overheads, coverage, grade→genus mapping, rate card, costing, FX and the
full transition plan. Every other sheet is locked and made entirely of live formulas
that reference the Inputs sheet, so you can change any input there and watch every
page + the Dashboard recalculate **without the app**.

Sheets:
  Summary       client-facing cover: headline price (INR + reporting cur), assumptions
  Inputs        the ONLY editable sheet — every application input
  Rate Cards    the full uploaded rate card (locked reference)
  1-2 Workload  per-severity hours + buffered role hours (formulas → Inputs)
  3 Patching    manual / tool effort → hours
  4 Activities  per-activity role-hour split
  5 Effort      base + contingency + overhead → assembled role hours
  6 FTE         productive hrs + coverage → FTE per role
  7 Rates       role → genus → rate (VLOOKUP on Inputs) × FX → INR
  Transition    phases + weekly utilisation → cost → commercial treatment
  8 Costing     resource cost + expenses + SLA → delivery → price (+ transition)
  Dashboard     executive / resource / effort / financial / transition summaries

Grey "App value" cells echo the tool's computed result for cross-checking. Currency
is INR (the app's base); the Summary also shows the reporting currency.
"""
import io
import os
from datetime import date

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import streamlit as st

from config.settings import (
    ALL_ROLES, OVERHEAD_ROLES, COVERAGE_APPLICABLE_ROLES, CATEGORY_SUBLABELS,
    DEFAULT_ROLE_BUFFER_PCT, COVERAGE_MODELS, APP_NAME, hx,
)
from modules.calculations.engine import filter_rate_card
from modules.state.session_manager import run_model

NAVY = hx("navy"); YEL = "FFF2CC"; LB = hx("tint"); GREY = hx("text_muted"); TEAL = hx("teal_dark")
_thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

S_SUM = "Summary"; S_IN = "Inputs"; S_RC = "Rate Cards"; S_WL = "1-2 Workload"
S_PA = "3 Patching"; S_AC = "4 Activities"; S_EF = "5 Effort"; S_FT = "6 FTE"
S_RT = "7 Rates"; S_TR = "Transition"; S_CO = "8 Costing"; S_DB = "Dashboard"

CAT_DISPLAY = [("alerts", "Monitoring Alerts"), ("service_requests", "Service Requests"),
               ("incidents", "Incidents"), ("changes", "Change Requests")]
ROLE_PCT_COL = {"L1": 3, "L2": 4, "L3": 5, "Architect": 6, "SDM": 7}


def _fill(c): return PatternFill("solid", fgColor=c)


def _aref(ws, row, col):
    return f"'{ws.title}'!${get_column_letter(col)}${row}"


def _hdr(ws, row, col, text):
    c = ws.cell(row, col); c.value = text; c.fill = _fill(NAVY)
    c.font = Font(color="FFFFFF", bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def _title(ws, row, col, text):
    ws.cell(row, col, text).font = Font(bold=True, color=NAVY, size=11)


def _lbl(ws, row, col, text, bold=False):
    ws.cell(row, col, text).font = Font(bold=bold, size=10)


def _edit(ws, row, col, value, fmt="#,##0.0"):
    """Editable INPUT cell (yellow, unlocked). The only kind of editable cell."""
    c = ws.cell(row, col); c.value = value; c.fill = _fill(YEL)
    c.number_format = fmt; c.border = BORDER
    c.protection = Protection(locked=False)
    return _aref(ws, row, col)


def _etext(ws, row, col, text):
    c = ws.cell(row, col); c.value = text; c.fill = _fill(YEL); c.border = BORDER
    c.protection = Protection(locked=False)
    return _aref(ws, row, col)


def _calc(ws, row, col, formula, fmt="#,##0.0"):
    """Locked formula cell (white)."""
    c = ws.cell(row, col); c.value = formula; c.number_format = fmt; c.border = BORDER
    return _aref(ws, row, col)


def _app(ws, row, col, value, fmt="#,##0.0"):
    c = ws.cell(row, col); c.value = round(float(value or 0), 2)
    c.number_format = fmt; c.font = Font(color=GREY, italic=True)


def generate_excel_model() -> bytes:
    model = st.session_state.get("_model") or run_model()
    ss = st.session_state
    wb = openpyxl.Workbook()
    ws_in = wb.active; ws_in.title = S_IN

    ws_in.column_dimensions["A"].width = 30
    for col in "BCD":
        ws_in.column_dimensions[col].width = 12
    for col in "EFGHIJKLMNOPQRSTUVWXYZ":
        ws_in.column_dimensions[col].width = 7
    ws_in["A1"] = APP_NAME; ws_in["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws_in["A2"] = ("THE ONLY EDITABLE SHEET. Every application input lives here (yellow). "
                   "Change any value and every page + the Dashboard recalculates. Currency: INR.")
    ws_in["A2"].font = Font(italic=True, size=9, color=GREY)
    ws_in["A3"] = f"Generated {date.today()}"; ws_in["A3"].font = Font(size=9, color=GREY)

    IN = {}
    r = 5

    def section(text):
        nonlocal r
        _title(ws_in, r, 1, text); r += 1

    def scalar(label, value, key, fmt="#,##0.0"):
        nonlocal r
        _lbl(ws_in, r, 1, label); IN[key] = _edit(ws_in, r, 2, value, fmt); r += 1

    def scalar_text(label, value, key):
        nonlocal r
        _lbl(ws_in, r, 1, label); IN[key] = _etext(ws_in, r, 2, value); r += 1

    # ── Engagement & assumptions ──────────────────────────────
    section("ENGAGEMENT")
    scalar_text("Customer / RFP name", str(ss.get("project_name") or ""), "project")
    scalar_text("Prepared by", str(ss.get("prepared_by") or ""), "prep")
    scalar_text("Delivery country", str(ss.get("delivery_country") or ""), "country")
    scalar_text("Delivery location", str(ss.get("delivery_location") or ""), "location")
    scalar_text("Reporting currency", str(ss.get("reporting_currency") or "INR"), "repcur")
    r += 1

    section("COVERAGE & FTE")
    scalar_text("Coverage model (8×5 / 24×7 / Custom …)", str(ss.get("coverage_model") or "8×5"), "covModel")
    scalar("Custom hours / day", float(ss.get("custom_hours_per_day", 8) or 8), "covHpd", "#,##0")
    scalar("Custom days / week", float(ss.get("custom_days_per_week", 5) or 5), "covDpw", "#,##0")
    scalar("Monthly working hours / FTE", float(ss.get("monthly_working_hours", 160) or 160), "monthly", "#,##0")
    scalar("Productive utilisation %", float(ss.get("productive_utilisation", 75) or 75), "util", "#,##0")
    scalar("FTE basis (1 = raw, 0 = rounded ⌈0.5⌉)", 1 if model.get("fte_basis") == "raw" else 0, "basisRaw", "0")
    scalar("Contingency %", float(ss.get("contingency_pct", 0) or 0), "cont", "#,##0")
    ovh = ss.get("overhead_pcts", {}) or {}
    scalar("Overhead % — Architect", float(ovh.get("Architect", 0) or 0), "ovhArchitect", "#,##0")
    scalar("Overhead % — SDM", float(ovh.get("SDM", 0) or 0), "ovhSDM", "#,##0")
    # coverage multiplier reference table (locked) + derived multiplier
    cov_first = r
    _lbl(ws_in, r, 4, "Coverage model"); _lbl(ws_in, r, 5, "Mult"); r += 1
    cov_tbl_first = r
    for m, cfg in COVERAGE_MODELS.items():
        if m == "Custom":
            continue
        ws_in.cell(r, 4, m); c = ws_in.cell(r, 5, round(cfg["multiplier"], 4)); c.number_format = "#,##0.0000"
        r += 1
    cov_tbl_last = r - 1
    COV_RANGE = f"'{S_IN}'!$D${cov_tbl_first}:$E${cov_tbl_last}"
    _lbl(ws_in, r, 1, "Coverage multiplier (derived)")
    IN["cov"] = _calc(ws_in, r, 2,
                      f'=IF({IN["covModel"]}="Custom",{IN["covHpd"]}*{IN["covDpw"]}/40,'
                      f'IFERROR(VLOOKUP({IN["covModel"]},{COV_RANGE},2,FALSE),1))', "#,##0.00")
    _app(ws_in, r, 3, model.get("coverage_multiplier", 1.0), "#,##0.00"); r += 2

    # ── Workload volumetrics (Steps 1-2) ──────────────────────
    section("MONTHLY WORKLOAD VOLUMETRICS (Steps 1–2)")
    wl_heads = ["Category", "Severity/Type", "Count", "Min", "L1 %", "L1 Buf%",
                "L2 %", "L2 Buf%", "L3 %", "L3 Buf%"]
    for i, h in enumerate(wl_heads, 1):
        _hdr(ws_in, r, i, h)
    r += 1
    WLI = {}
    for cat_key, cat_label in CAT_DISPLAY:
        cat = ss.get(cat_key, {}) or {}
        WLI[cat_key] = {}
        for label in CATEGORY_SUBLABELS[cat_key]:
            row = cat.get(label, {})
            _lbl(ws_in, r, 1, cat_label); _lbl(ws_in, r, 2, label)
            WLI[cat_key][label] = {
                "count": _edit(ws_in, r, 3, int(row.get("count", 0) or 0), "#,##0"),
                "min":   _edit(ws_in, r, 4, float(row.get("minutes", 0) or 0), "#,##0"),
                "l1":    _edit(ws_in, r, 5, float(row.get("L1_pct", 0) or 0), "#,##0"),
                "l1b":   _edit(ws_in, r, 6, float(row.get("L1_buffer", DEFAULT_ROLE_BUFFER_PCT)), "#,##0"),
                "l2":    _edit(ws_in, r, 7, float(row.get("L2_pct", 0) or 0), "#,##0"),
                "l2b":   _edit(ws_in, r, 8, float(row.get("L2_buffer", DEFAULT_ROLE_BUFFER_PCT)), "#,##0"),
                "l3":    _edit(ws_in, r, 9, float(row.get("L3_pct", 0) or 0), "#,##0"),
                "l3b":   _edit(ws_in, r, 10, float(row.get("L3_buffer", DEFAULT_ROLE_BUFFER_PCT)), "#,##0"),
            }
            r += 1
    r += 1

    # ── Patching (Step 3) ─────────────────────────────────────
    section("PATCHING (Step 3)")
    scalar("Patching included (1 = yes)", 1 if ss.get("patching_included") == "Yes" else 0, "patchInc", "0")
    scalar("Number of servers", int(ss.get("num_servers", 0) or 0), "servers", "#,##0")
    scalar("Tool-based (1 = tool, 0 = manual)", 1 if ss.get("patching_method") == "Tool-Based" else 0, "patchTool", "0")
    scalar("Manual effort (min/server)", float(ss.get("manual_effort_per_server", 45) or 45), "manMin", "#,##0")
    scalar("Tool effort (min/failed server)", float(ss.get("auto_effort_per_server", 30) or 30), "autoMin", "#,##0")
    scalar("Tool error/failure rate %", float(ss.get("patch_error_rate", 0) or 0), "errRate", "#,##0")
    scalar_text("Patching assigned to role", ss.get("patching_role", "L2") or "L2", "patchRole")
    r += 1

    # ── Additional activities (Step 4) ────────────────────────
    section("ADDITIONAL ACTIVITIES (Step 4)")
    for i, h in enumerate(["Activity", "Monthly Hrs", "L1 %", "L2 %", "L3 %", "Arch %", "SDM %"], 1):
        _hdr(ws_in, r, i, h)
    r += 1
    ACI = []
    acts = ss.get("additional_activities", []) or []
    if not acts:
        acts = [{"name": "—", "hours": 0.0, "dist": {}}]
    for act in acts:
        dist = act.get("dist", {}) or {}
        rec = {"name": _etext(ws_in, r, 1, str(act.get("name", "Activity"))),
               "hours": _edit(ws_in, r, 2, float(act.get("hours", 0) or 0), "#,##0.0")}
        for role, col in ROLE_PCT_COL.items():
            rec[role] = _edit(ws_in, r, col, float(dist.get(role, 0) or 0), "#,##0")
        ACI.append(rec); r += 1
    r += 1

    # ── Grade mapping (Step 7) ────────────────────────────────
    section("GRADE MAPPING — role → genus (Step 7)")
    GENUS = {}
    rc = model["resource_costs"]
    for role in ALL_ROLES:
        _lbl(ws_in, r, 1, role)
        GENUS[role] = _etext(ws_in, r, 2, str(rc.get(role, {}).get("genus") or ""))
        r += 1
    r += 1

    # ── Scoped rate card (Step 7 source) ──────────────────────
    section("SCOPED RATE CARD (genus → hourly rate)")
    df = ss.get("rate_card_df")
    country = ss.get("delivery_country"); location = ss.get("delivery_location")
    for i, h in enumerate(["Genus", "Hourly Rate", "Currency"], 1):
        _hdr(ws_in, r, i, h)
    r += 1
    scoped_first = r
    if df is not None and len(df) > 0:
        scoped = filter_rate_card(df, country, location).drop_duplicates(subset=["genus"], keep="first")
        for _, row in scoped.iterrows():
            _etext(ws_in, r, 1, str(row.get("genus", "")))
            _edit(ws_in, r, 2, float(row.get("hourly rate", 0) or 0), "#,##0.00")
            _etext(ws_in, r, 3, str(row.get("rate currency", "INR") or "INR").upper().strip())
            r += 1
    if r == scoped_first:
        _etext(ws_in, r, 1, "—"); _edit(ws_in, r, 2, 0, "#,##0.00"); _etext(ws_in, r, 3, "INR"); r += 1
    scoped_last = r - 1
    SCOPED_RANGE = f"'{S_IN}'!$A${scoped_first}:$C${scoped_last}"
    r += 1

    # ── Costing (Step 8) ──────────────────────────────────────
    section("COSTING (Step 8)")
    scalar("Target gross margin %", float(model["price_result"].get("margin_pct", 0) or 0), "margin", "#,##0")
    scalar("SLA provision included (1 = yes)", 1 if ss.get("sla_provision_included") == "Yes" else 0, "slaInc", "0")
    scalar("SLA provision %", float(ss.get("sla_provision_pct", 0) or 0), "slaPct", "#,##0.0")
    r += 1
    _lbl(ws_in, r, 1, "FX rates — 1 unit = X INR", bold=True); r += 1
    _hdr(ws_in, r, 1, "Currency"); _hdr(ws_in, r, 2, "INR per unit"); r += 1
    fx = {"INR": 1.0}
    for k, v in (ss.get("exchange_rates", {}) or {}).items():
        if v:
            fx[str(k).upper().strip()] = float(v)
    fx_first = r
    for cur, rate in fx.items():
        _etext(ws_in, r, 1, cur); _edit(ws_in, r, 2, rate, "#,##0.0000"); r += 1
    fx_last = r - 1
    FX_RANGE = f"'{S_IN}'!$A${fx_first}:$B${fx_last}"
    r += 1
    _lbl(ws_in, r, 1, "Additional cost items (INR)", bold=True); r += 1
    _hdr(ws_in, r, 1, "Item"); _hdr(ws_in, r, 2, "Monthly cost"); r += 1
    add_first = r
    add_costs = ss.get("additional_costs", []) or []
    if not add_costs:
        _etext(ws_in, r, 1, "—"); _edit(ws_in, r, 2, 0, "#,##0"); r += 1
    for row in add_costs:
        _etext(ws_in, r, 1, str(row.get("name", "Cost")))
        _edit(ws_in, r, 2, float(row.get("cost", 0) or 0), "#,##0"); r += 1
    add_last = r - 1
    _lbl(ws_in, r, 1, "Total additional expenses", bold=True)
    IN["expenses"] = _calc(ws_in, r, 2, f"=SUM(B{add_first}:B{add_last})", "#,##0"); r += 2

    # ── Transition plan (Step 8) ──────────────────────────────
    section("TRANSITION & ONBOARDING (Step 8)")
    tp = ss.get("transition_planner", {}) or {}
    mtrans = model.get("transition", {}) or {}
    tr_enabled = bool(tp.get("enabled"))
    total_weeks = int(tp.get("total_weeks", 0) or 0) if tr_enabled else 0
    phases = tp.get("phases", []) or []
    resources = tp.get("resources", []) or []
    allocation = tp.get("allocation", {}) or {}
    scalar("Transition included (1 = yes)", 1 if tr_enabled else 0, "trInc", "0")
    scalar("Total transition weeks", total_weeks, "trWeeks", "#,##0")
    scalar("Weekly hours / full-capacity", float(mtrans.get("weekly_hours", 40) or 40), "trWH", "#,##0")
    scalar_text("Treatment (recurring/one_time/absorb)", str(tp.get("treatment", "recurring")), "trTreat")
    scalar("Amortisation months", int(tp.get("amortisation_months", 12) or 12), "trMonths", "#,##0")
    TRI = {"phases": [], "resources": [], "alloc": {}}
    if tr_enabled and total_weeks > 0 and resources:
        # phases
        _lbl(ws_in, r, 1, "Phase", bold=True); _lbl(ws_in, r, 2, "Weeks", bold=True); r += 1
        for ph in phases:
            nm = _etext(ws_in, r, 1, str(ph.get("name", "")))
            wk = _edit(ws_in, r, 2, int(ph.get("weeks", 0) or 0), "#,##0")
            TRI["phases"].append({"name": nm, "weeks": wk, "name_val": str(ph.get("name", "")),
                                  "weeks_val": int(ph.get("weeks", 0) or 0)})
            r += 1
        # weekly utilisation grid: Resource | Role | Count | W1..WN
        _hdr(ws_in, r, 1, "Resource"); _hdr(ws_in, r, 2, "Role"); _hdr(ws_in, r, 3, "Count")
        for w in range(1, total_weeks + 1):
            _hdr(ws_in, r, 3 + w, f"W{w}")
        r += 1
        seen = {}
        util_first = r
        for res in resources:
            rid = str(res.get("id", "")); role = res.get("role")
            seen[role] = seen.get(role, 0) + 1
            _etext(ws_in, r, 1, f"{role} #{seen[role]}")
            role_ref = _etext(ws_in, r, 2, str(role))
            cnt_ref = _edit(ws_in, r, 3, float(res.get("count", 0) or 0), "#,##0")
            alloc = allocation.get(rid) or allocation.get(res.get("id")) or {}
            wk_refs = {}
            for w in range(1, total_weeks + 1):
                util = float(alloc.get(str(w), alloc.get(w, 0)) or 0)
                wk_refs[w] = _edit(ws_in, r, 3 + w, util, "0.00")
            TRI["resources"].append({"id": rid, "role": role, "role_ref": role_ref,
                                     "count": cnt_ref, "weeks": wk_refs})
            r += 1
        util_last = r - 1
        dv = DataValidation(type="list", formula1='"0,0.25,0.5,1"', allow_blank=True)
        ws_in.add_data_validation(dv)
        dv.add(f"{get_column_letter(4)}{util_first}:{get_column_letter(3 + total_weeks)}{util_last}")
    r += 1

    # ════════════════════════════════════════════════════════════════════════
    # RATE CARDS  (full uploaded card — locked reference)
    # ════════════════════════════════════════════════════════════════════════
    ws_rc = wb.create_sheet(S_RC)
    for col, w in (("A", 16), ("B", 16), ("C", 20), ("D", 14), ("E", 10)):
        ws_rc.column_dimensions[col].width = w
    _title(ws_rc, 1, 1, f"Full uploaded rate card — reference ({country or 'All'}"
                        + (f" / {location}" if location else "") + ")")
    ws_rc.cell(2, 1, "The scoped genus→rate lookup used by 7 Rates is on the Inputs sheet "
                     "(editable). This is the full source card.").font = Font(italic=True, size=9, color=GREY)
    rr = 4
    for i, h in enumerate(["Country", "Location", "Genus", "Hourly Rate", "Currency"], 1):
        _hdr(ws_rc, rr, i, h)
    rr += 1
    if df is not None and len(df) > 0:
        for _, row in df.iterrows():
            ws_rc.cell(rr, 1, str(row.get("country", "")))
            ws_rc.cell(rr, 2, str(row.get("location", "")))
            ws_rc.cell(rr, 3, str(row.get("genus", "")))
            c = ws_rc.cell(rr, 4, float(row.get("hourly rate", 0) or 0)); c.number_format = "#,##0.00"
            ws_rc.cell(rr, 5, str(row.get("rate currency", "INR") or "INR").upper().strip())
            rr += 1

    # ════════════════════════════════════════════════════════════════════════
    # 1-2 WORKLOAD  (formulas referencing Inputs)
    # ════════════════════════════════════════════════════════════════════════
    ws_wl = wb.create_sheet(S_WL)
    headers = ["Category", "Severity/Type", "Count", "Min", "L1 %", "L1 Buf%",
               "L2 %", "L2 Buf%", "L3 %", "L3 Buf%", "Cat Hrs", "L1 Hrs", "L2 Hrs", "L3 Hrs"]
    ws_wl.column_dimensions["A"].width = 18; ws_wl.column_dimensions["B"].width = 14
    for col in "CDEFGHIJKLMN":
        ws_wl.column_dimensions[col].width = 9
    _title(ws_wl, 1, 1, "Workload → effort (Steps 1–2) — inputs mirrored from the Inputs sheet")
    hr = 3
    for i, h in enumerate(headers, 1):
        _hdr(ws_wl, hr, i, h)
    r = hr + 1
    WL_cat_hrs = {}; WL_vol = {}
    all_first = r
    for cat_key, cat_label in CAT_DISPLAY:
        c_first = r
        for label in CATEGORY_SUBLABELS[cat_key]:
            ref = WLI[cat_key][label]
            _lbl(ws_wl, r, 1, cat_label); _lbl(ws_wl, r, 2, label)
            _calc(ws_wl, r, 3, f"={ref['count']}", "#,##0")
            _calc(ws_wl, r, 4, f"={ref['min']}", "#,##0")
            _calc(ws_wl, r, 5, f"={ref['l1']}", "#,##0")
            _calc(ws_wl, r, 6, f"={ref['l1b']}", "#,##0")
            _calc(ws_wl, r, 7, f"={ref['l2']}", "#,##0")
            _calc(ws_wl, r, 8, f"={ref['l2b']}", "#,##0")
            _calc(ws_wl, r, 9, f"={ref['l3']}", "#,##0")
            _calc(ws_wl, r, 10, f"={ref['l3b']}", "#,##0")
            _calc(ws_wl, r, 11, f"=C{r}*D{r}/60")
            _calc(ws_wl, r, 12, f"=C{r}*D{r}/60*E{r}/100*(1+F{r}/100)")
            _calc(ws_wl, r, 13, f"=C{r}*D{r}/60*G{r}/100*(1+H{r}/100)")
            _calc(ws_wl, r, 14, f"=C{r}*D{r}/60*I{r}/100*(1+J{r}/100)")
            r += 1
        c_last = r - 1
        _lbl(ws_wl, r, 2, f"{cat_label} subtotal", bold=True)
        WL_vol[cat_key] = _calc(ws_wl, r, 3, f"=SUM(C{c_first}:C{c_last})", "#,##0")
        WL_cat_hrs[cat_key] = _calc(ws_wl, r, 11, f"=SUM(K{c_first}:K{c_last})")
        ws_wl.cell(r, 1).fill = _fill(LB); r += 1
    all_last = r - 1
    _lbl(ws_wl, r, 2, "TOTAL buffered role hours", bold=True)
    WL_role = {}
    for role, col in (("L1", 12), ("L2", 13), ("L3", 14)):
        L = get_column_letter(col)
        WL_role[role] = _calc(ws_wl, r, col, f"=SUM({L}{all_first}:{L}{all_last})")
    for c in range(2, 15):
        ws_wl.cell(r, c).fill = _fill(TEAL); ws_wl.cell(r, c).font = Font(bold=True, color="FFFFFF")

    # ════════════════════════════════════════════════════════════════════════
    # 3 PATCHING
    # ════════════════════════════════════════════════════════════════════════
    ws_pa = wb.create_sheet(S_PA)
    ws_pa.column_dimensions["A"].width = 40; ws_pa.column_dimensions["B"].width = 16
    _title(ws_pa, 1, 1, "Patching effort (Step 3)")
    _lbl(ws_pa, 3, 1, "Failed servers (tool mode) = ROUND(servers × error%)")
    _calc(ws_pa, 3, 2, f"=IF({IN['patchTool']}=1,ROUND({IN['servers']}*{IN['errRate']}/100,0),0)", "#,##0")
    _lbl(ws_pa, 4, 1, "Patching hours / month", bold=True)
    PATCH_HRS = _calc(ws_pa, 4, 2,
                      f"=IF({IN['patchInc']}=1,IF({IN['patchTool']}=1,B3*{IN['autoMin']}/60,"
                      f"{IN['servers']}*{IN['manMin']}/60),0)")
    _app(ws_pa, 4, 3, model["effort_sources"].get("Patching", 0))
    _lbl(ws_pa, 5, 1, "Assigned to role"); _calc(ws_pa, 5, 2, f"={IN['patchRole']}", "General")

    # ════════════════════════════════════════════════════════════════════════
    # 4 ACTIVITIES
    # ════════════════════════════════════════════════════════════════════════
    ws_ac = wb.create_sheet(S_AC)
    ws_ac.column_dimensions["A"].width = 30
    for col in "BCDEFGH":
        ws_ac.column_dimensions[col].width = 10
    _title(ws_ac, 1, 1, "Additional activities (Step 4)")
    for i, h in enumerate(["Activity", "Monthly Hrs", "L1 %", "L2 %", "L3 %", "Arch %", "SDM %"], 1):
        _hdr(ws_ac, 3, i, h)
    r = 4
    ac_first = r
    for rec in ACI:
        _calc(ws_ac, r, 1, f"={rec['name']}", "General")
        _calc(ws_ac, r, 2, f"={rec['hours']}", "#,##0.0")
        for role, col in ROLE_PCT_COL.items():
            _calc(ws_ac, r, col, f"={rec[role]}", "#,##0")
        r += 1
    ac_last = r - 1
    _lbl(ws_ac, r, 1, "Total hours / role hours", bold=True)
    AC_total = _calc(ws_ac, r, 2, f"=SUM(B{ac_first}:B{ac_last})")
    AC_add = {}
    for role, col in ROLE_PCT_COL.items():
        L = get_column_letter(col)
        AC_add[role] = (f"(SUMPRODUCT('{S_AC}'!$B${ac_first}:$B${ac_last},"
                        f"'{S_AC}'!{L}${ac_first}:{L}${ac_last})/100)")
    for c in range(1, 9):
        ws_ac.cell(r, c).fill = _fill(LB)

    # ════════════════════════════════════════════════════════════════════════
    # 5 EFFORT
    # ════════════════════════════════════════════════════════════════════════
    ws_ef = wb.create_sheet(S_EF)
    ws_ef.column_dimensions["A"].width = 30
    for col in "BCD":
        ws_ef.column_dimensions[col].width = 14
    _title(ws_ef, 1, 1, "Effort summary (Step 5)")
    _hdr(ws_ef, 3, 1, "Source"); _hdr(ws_ef, 3, 2, "Hours"); _hdr(ws_ef, 3, 3, "App value")
    r = 4
    EF = {}
    es = model["effort_sources"]

    def effort_row(label, formula, app_val, key=None, fmt="#,##0.0"):
        nonlocal r
        _lbl(ws_ef, r, 1, label); ref = _calc(ws_ef, r, 2, formula, fmt)
        _app(ws_ef, r, 3, app_val, fmt)
        if key:
            EF[key] = ref
        r += 1
        return ref

    effort_row("Monitoring Alerts", f"={WL_cat_hrs['alerts']}", es.get("Monitoring Alerts", 0))
    effort_row("Service Requests", f"={WL_cat_hrs['service_requests']}", es.get("Service Requests", 0))
    effort_row("Incidents", f"={WL_cat_hrs['incidents']}", es.get("Incidents", 0))
    effort_row("Change Requests", f"={WL_cat_hrs['changes']}", es.get("Change Requests", 0))
    effort_row("Patching", f"={PATCH_HRS}", es.get("Patching", 0))
    effort_row("Additional Activities", f"={AC_total}", es.get("Additional Activities", 0))
    EF["base"] = effort_row("Base Effort",
                            f"={WL_cat_hrs['alerts']}+{WL_cat_hrs['service_requests']}+{WL_cat_hrs['incidents']}"
                            f"+{WL_cat_hrs['changes']}+{PATCH_HRS}+{AC_total}",
                            model["base_effort"], key="base")
    EF["contHrs"] = effort_row("Contingency", f"={EF['base']}*{IN['cont']}/100",
                               es.get("Contingency", 0), key="contHrs")
    EF["total"] = effort_row("Total operational effort", f"={EF['base']}+{EF['contHrs']}",
                             model["total_effort"], key="total")

    r += 1
    _title(ws_ef, r, 1, "Assembled role hours"); r += 1
    _hdr(ws_ef, r, 1, "Role"); _hdr(ws_ef, r, 2, "Hours"); _hdr(ws_ef, r, 3, "App value"); r += 1
    EF_role = {}
    cont_mult = f"(1+{IN['cont']}/100)"
    ovh_ref = {"Architect": IN["ovhArchitect"], "SDM": IN["ovhSDM"]}
    for role in ALL_ROLES:
        ticket = WL_role.get(role)
        base_terms = []
        if ticket:
            base_terms.append(ticket)
        base_terms.append(AC_add[role])
        base_terms.append(f'IF({IN["patchRole"]}="{role}",{PATCH_HRS},0)')
        base_expr = "+".join(base_terms)
        ovh_term = f"+{EF['total']}*{ovh_ref[role]}/100" if role in OVERHEAD_ROLES else ""
        _lbl(ws_ef, r, 1, role)
        EF_role[role] = _calc(ws_ef, r, 2, f"=({base_expr})*{cont_mult}{ovh_term}")
        _app(ws_ef, r, 3, model["role_hours"].get(role, 0))
        r += 1

    # ════════════════════════════════════════════════════════════════════════
    # 6 FTE
    # ════════════════════════════════════════════════════════════════════════
    ws_ft = wb.create_sheet(S_FT)
    ws_ft.column_dimensions["A"].width = 14
    for col in "BCDEFG":
        ws_ft.column_dimensions[col].width = 12
    _title(ws_ft, 1, 1, "FTE calculation (Step 6)")
    _lbl(ws_ft, 3, 1, "Productive hrs / FTE", bold=True)
    PROD = _calc(ws_ft, 3, 2, f"={IN['monthly']}*{IN['util']}/100", "#,##0.0")
    _app(ws_ft, 3, 3, model.get("productive_hours", 0))
    hr = 5
    for i, h in enumerate(["Role", "Hours", "Raw FTE (adj)", "Final FTE ⌈0.5⌉", "FTE used", "App FTE used"], 1):
        _hdr(ws_ft, hr, i, h)
    r = hr + 1
    FT_fte = {}
    fte_key = "raw_fte" if model.get("fte_basis") == "raw" else "final_fte"
    for role in ALL_ROLES:
        _lbl(ws_ft, r, 1, role)
        hcell = _calc(ws_ft, r, 2, f"={EF_role[role]}")
        covf = f"*{IN['cov']}" if role in COVERAGE_APPLICABLE_ROLES else ""
        raw = _calc(ws_ft, r, 3, f"=IF({PROD}>0,{hcell}/{PROD}{covf},0)", "#,##0.000")
        final = _calc(ws_ft, r, 4, f"=IF({hcell}>0,MAX(CEILING({raw},0.5),0.5),0)", "#,##0.00")
        FT_fte[role] = _calc(ws_ft, r, 5, f"=IF({IN['basisRaw']}=1,{raw},{final})", "#,##0.00")
        _app(ws_ft, r, 6, model["fte_result"].get(role, {}).get(fte_key, 0), "#,##0.00")
        r += 1
    _lbl(ws_ft, r, 1, "TOTAL", bold=True)
    FT_total = _calc(ws_ft, r, 5, "=" + "+".join(FT_fte[x] for x in ALL_ROLES), "#,##0.00")
    _app(ws_ft, r, 6, model["total_fte"], "#,##0.00")

    # ════════════════════════════════════════════════════════════════════════
    # 7 RATES
    # ════════════════════════════════════════════════════════════════════════
    ws_rt = wb.create_sheet(S_RT)
    ws_rt.column_dimensions["A"].width = 12; ws_rt.column_dimensions["B"].width = 20
    for col in "CDEF":
        ws_rt.column_dimensions[col].width = 14
    _title(ws_rt, 1, 1, "Role → genus → rate (Step 7)")
    ws_rt.cell(2, 1, "Genus is looked up on the Inputs scoped rate card → × FX → INR.").font = \
        Font(italic=True, size=9, color=GREY)
    hr = 4
    for i, h in enumerate(["Role", "Genus", "Rate (native)", "Currency", "Rate (INR)", "App INR"], 1):
        _hdr(ws_rt, hr, i, h)
    r = hr + 1
    RT_rate = {}
    for role in ALL_ROLES:
        _lbl(ws_rt, r, 1, role)
        g = _calc(ws_rt, r, 2, f"={GENUS[role]}", "General")
        native = _calc(ws_rt, r, 3, f"=IFERROR(VLOOKUP({g},{SCOPED_RANGE},2,FALSE),0)", "#,##0.00")
        cur = _calc(ws_rt, r, 4, f'=IFERROR(VLOOKUP({g},{SCOPED_RANGE},3,FALSE),"INR")', "General")
        RT_rate[role] = _calc(ws_rt, r, 5,
                              f"={native}*IFERROR(VLOOKUP({cur},{FX_RANGE},2,FALSE),1)", "#,##0")
        _app(ws_rt, r, 6, rc.get(role, {}).get("rate_inr", 0), "#,##0")
        r += 1

    # ════════════════════════════════════════════════════════════════════════
    # TRANSITION  (formulas referencing Inputs)
    # ════════════════════════════════════════════════════════════════════════
    ws_tr = wb.create_sheet(S_TR)
    for col, w in (("A", 18), ("B", 12), ("C", 10), ("D", 12)):
        ws_tr.column_dimensions[col].width = w
    _title(ws_tr, 1, 1, "Transition & Onboarding Planner")
    ws_tr.cell(2, 1, "Weekly cost = count × utilisation × weekly hours × role rate. "
                     "Inputs are on the Inputs sheet.").font = Font(italic=True, size=9, color=GREY)
    TR_phase_refs = {}
    if not (tr_enabled and total_weeks > 0 and resources):
        _lbl(ws_tr, 4, 1, "Transition & Onboarding is not included in this estimate.", bold=True)
        ws_tr.cell(5, 2, 0).number_format = "#,##0"
        _zero = _aref(ws_tr, 5, 2)
        TR_total = TR_monthly = TR_onetime = TR_absorb = TR_net = _zero
    else:
        # weekly cost grid: Resource | Role | Count | Rate | W1..WN | Total Hrs | Cost
        wk_col0 = 5
        hr = 4
        _hdr(ws_tr, hr, 1, "Resource"); _hdr(ws_tr, hr, 2, "Role")
        _hdr(ws_tr, hr, 3, "Count"); _hdr(ws_tr, hr, 4, "Rate")
        for w in range(1, total_weeks + 1):
            c = wk_col0 + (w - 1)
            _hdr(ws_tr, hr, c, f"W{w}"); ws_tr.column_dimensions[get_column_letter(c)].width = 6
        hrs_col = wk_col0 + total_weeks; cost_col = hrs_col + 1
        _hdr(ws_tr, hr, hrs_col, "Total Hrs"); _hdr(ws_tr, hr, cost_col, "Cost INR")
        ws_tr.column_dimensions[get_column_letter(hrs_col)].width = 11
        ws_tr.column_dimensions[get_column_letter(cost_col)].width = 14
        r = hr + 1
        res_first = r
        wk_first_L = get_column_letter(wk_col0); wk_last_L = get_column_letter(wk_col0 + total_weeks - 1)
        hrs_L = get_column_letter(hrs_col); cost_L = get_column_letter(cost_col)
        seen = {}
        for res in TRI["resources"]:
            role = res["role"]; seen[role] = seen.get(role, 0) + 1
            _calc(ws_tr, r, 1, f'="{role} #{seen[role]}"', "General")
            _calc(ws_tr, r, 2, f"={res['role_ref']}", "General")
            _calc(ws_tr, r, 3, f"={res['count']}", "#,##0")
            rr = RT_rate.get(role)
            _calc(ws_tr, r, 4, f"={rr}" if rr else "=0", "#,##0")
            for w in range(1, total_weeks + 1):
                _calc(ws_tr, r, wk_col0 + (w - 1), f"={res['weeks'][w]}", "0.00")
            _calc(ws_tr, r, hrs_col, f"=C{r}*{IN['trWH']}*SUM({wk_first_L}{r}:{wk_last_L}{r})", "#,##0.0")
            _calc(ws_tr, r, cost_col, f"={hrs_L}{r}*D{r}", "#,##0")
            r += 1
        res_last = r - 1
        _lbl(ws_tr, r, 1, "TOTAL", bold=True)
        TR_total = _calc(ws_tr, r, cost_col, f"=SUM({cost_L}{res_first}:{cost_L}{res_last})", "#,##0")
        _app(ws_tr, r, cost_col + 1, mtrans.get("total", 0)); r += 2

        # cost by phase (contiguous week columns, mapped from phase weeks on Inputs)
        _title(ws_tr, r, 1, "Cost by phase"); r += 1
        _hdr(ws_tr, r, 1, "Phase"); _hdr(ws_tr, r, 2, "Cost INR"); _hdr(ws_tr, r, 3, "App value"); r += 1
        cnt_rng = f"$C${res_first}:$C${res_last}"; rate_rng = f"$D${res_first}:$D${res_last}"
        wk_cursor = 1
        for ph in TRI["phases"]:
            nm_val = ph["name_val"]; wks = ph["weeks_val"]
            start = wk_cursor; end = wk_cursor + wks - 1; wk_cursor = end + 1
            _calc(ws_tr, r, 1, f"={ph['name']}", "General")
            if end >= start and start >= 1 and end <= total_weeks:
                cols = [get_column_letter(wk_col0 + (w - 1)) for w in range(start, end + 1)]
                util_sum = "+".join(f"{c}{res_first}:{c}{res_last}" for c in cols)
                ref = _calc(ws_tr, r, 2, f"={IN['trWH']}*SUMPRODUCT({cnt_rng},{rate_rng},({util_sum}))", "#,##0")
            else:
                ref = _calc(ws_tr, r, 2, "=0", "#,##0")
            TR_phase_refs[nm_val] = ref
            _app(ws_tr, r, 3, mtrans.get("per_phase", {}).get(nm_val, 0)); r += 1
        r += 1

        _title(ws_tr, r, 1, "Commercial treatment"); r += 1
        _hdr(ws_tr, r, 1, "Item"); _hdr(ws_tr, r, 2, "INR"); _hdr(ws_tr, r, 3, "App value"); r += 1
        _lbl(ws_tr, r, 1, "Total transition cost")
        _calc(ws_tr, r, 2, f"={TR_total}", "#,##0"); _app(ws_tr, r, 3, mtrans.get("total", 0)); r += 1
        _lbl(ws_tr, r, 1, "Monthly recurring (÷ months)")
        TR_monthly = _calc(ws_tr, r, 2, f'=IF({IN["trTreat"]}="recurring",{TR_total}/MAX({IN["trMonths"]},1),0)', "#,##0")
        _app(ws_tr, r, 3, mtrans.get("monthly_recurring", 0)); r += 1
        _lbl(ws_tr, r, 1, "One-time fee")
        TR_onetime = _calc(ws_tr, r, 2, f'=IF({IN["trTreat"]}="one_time",{TR_total},0)', "#,##0")
        _app(ws_tr, r, 3, mtrans.get("one_time_fee", 0)); r += 1
        _lbl(ws_tr, r, 1, "Absorbed (discount)")
        TR_absorb = _calc(ws_tr, r, 2, f'=IF({IN["trTreat"]}="absorb",{TR_total},0)', "#,##0")
        _app(ws_tr, r, 3, mtrans.get("absorbed", 0)); r += 1
        _lbl(ws_tr, r, 1, "Net charged", bold=True)
        TR_net = _calc(ws_tr, r, 2, f"={TR_total}-{TR_absorb}", "#,##0")
        _app(ws_tr, r, 3, mtrans.get("net_charged", 0)); r += 1

    # ════════════════════════════════════════════════════════════════════════
    # 8 COSTING
    # ════════════════════════════════════════════════════════════════════════
    ws_co = wb.create_sheet(S_CO)
    ws_co.column_dimensions["A"].width = 30
    for col in "BCDE":
        ws_co.column_dimensions[col].width = 14
    _title(ws_co, 1, 1, "Costing & price (Step 8)")
    hr = 3
    for i, h in enumerate(["Role", "FTE", "Billed Hrs", "Rate INR", "Cost INR"], 1):
        _hdr(ws_co, hr, i, h)
    r = hr + 1
    CO_cost = {}
    for role in ALL_ROLES:
        _lbl(ws_co, r, 1, role)
        _calc(ws_co, r, 2, f"={FT_fte[role]}", "#,##0.00")
        billed = _calc(ws_co, r, 3, f"={FT_fte[role]}*{IN['monthly']}", "#,##0")
        _calc(ws_co, r, 4, f"={RT_rate[role]}", "#,##0")
        CO_cost[role] = _calc(ws_co, r, 5,
                              f"=IF(AND({FT_fte[role]}>0,{RT_rate[role]}>0),{billed}*{RT_rate[role]},0)", "#,##0")
        r += 1
    _lbl(ws_co, r, 1, "Total resource cost", bold=True)
    CO_res = _calc(ws_co, r, 5, "=" + "+".join(CO_cost[x] for x in ALL_ROLES), "#,##0")
    _app(ws_co, r, 4, model["total_resource_cost"], "#,##0"); r += 2

    _title(ws_co, r, 1, "Delivery → price"); r += 1
    _hdr(ws_co, r, 1, "Component"); _hdr(ws_co, r, 2, "INR"); _hdr(ws_co, r, 3, "App value"); r += 1

    def cost_row(label, formula, app_val):
        nonlocal r
        _lbl(ws_co, r, 1, label); ref = _calc(ws_co, r, 2, formula, "#,##0")
        _app(ws_co, r, 3, app_val, "#,##0")
        r += 1
        return ref
    CO = {}
    CO["res"] = cost_row("Total resource cost", f"={CO_res}", model["cost_result"]["resource_cost"])
    CO["exp"] = cost_row("Additional expenses", f"={IN['expenses']}", model["cost_result"]["additional_expenses"])
    sub = cost_row("Subtotal before SLA", f"={CO['res']}+{CO['exp']}", model["cost_result"]["subtotal_before_sla"])
    CO["sla"] = cost_row("SLA provision", f"={sub}*IF({IN['slaInc']}=1,{IN['slaPct']},0)/100",
                         model["cost_result"]["sla_provision"])
    CO["deliver"] = cost_row("TOTAL DELIVERY COST", f"={sub}+{CO['sla']}",
                             model["cost_result"]["total_delivery_cost"])
    CO["sell"] = cost_row("Monthly selling price", f"={CO['deliver']}/(1-{IN['margin']}/100)",
                          model["price_result"]["selling_price"])
    CO["profit"] = cost_row("Gross profit", f"={CO['sell']}-{CO['deliver']}",
                            model["price_result"]["gross_profit"])
    cost_row("Transition — total cost", f"={TR_total}", mtrans.get("total", 0))
    cost_row("Transition — monthly recurring", f"={TR_monthly}", mtrans.get("monthly_recurring", 0))
    cost_row("Transition — one-time fee", f"={TR_onetime}", mtrans.get("one_time_fee", 0))
    cost_row("Transition — absorbed (discount)", f"=-{TR_absorb}", -float(mtrans.get("absorbed", 0) or 0))
    CO["monthly_final"] = cost_row("MONTHLY PRICE incl. transition", f"={CO['sell']}+{TR_monthly}",
                                   model.get("monthly_price_with_transition", 0))

    # ════════════════════════════════════════════════════════════════════════
    # DASHBOARD
    # ════════════════════════════════════════════════════════════════════════
    ws_db = wb.create_sheet(S_DB)
    ws_db.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ws_db.column_dimensions[col].width = 15
    _title(ws_db, 1, 1, "Dashboard — fully dynamic (recalculates with any input change)")
    r = 3
    _title(ws_db, r, 1, "Executive Summary"); r += 1
    _hdr(ws_db, r, 1, "Metric"); _hdr(ws_db, r, 2, "Value"); _hdr(ws_db, r, 3, "App value"); r += 1
    for label, formula, app_val, fmt in [
        ("Total Effort (hrs)", f"={EF['total']}", model["total_effort"], "#,##0.0"),
        ("Total FTE", f"={FT_total}", model["total_fte"], "#,##0.00"),
        ("Delivery Cost (INR)", f"={CO['deliver']}", model["cost_result"]["total_delivery_cost"], "#,##0"),
        ("Selling Price (INR)", f"={CO['sell']}", model["price_result"]["selling_price"], "#,##0"),
        ("Gross Margin %", f"={IN['margin']}", model["price_result"]["margin_pct"], "#,##0.0"),
        ("Base Effort (hrs)", f"={EF['base']}", model["base_effort"], "#,##0.0"),
        ("Contingency (hrs)", f"={EF['contHrs']}", model["effort_sources"].get("Contingency", 0), "#,##0.0"),
        ("Transition Total (INR)", f"={TR_total}", mtrans.get("total", 0), "#,##0"),
        ("Monthly Price incl. Transition", f"={CO['monthly_final']}",
         model.get("monthly_price_with_transition", 0), "#,##0"),
    ]:
        _lbl(ws_db, r, 1, label); _calc(ws_db, r, 2, formula, fmt); _app(ws_db, r, 3, app_val, fmt); r += 1

    r += 1
    _title(ws_db, r, 1, "Resource Cost Summary"); r += 1
    for i, h in enumerate(["Role", "Genus", "FTE", "Billed Hrs", "Rate INR", "Cost INR"], 1):
        _hdr(ws_db, r, i, h)
    r += 1
    for role in ALL_ROLES:
        _lbl(ws_db, r, 1, role)
        _calc(ws_db, r, 2, f"={GENUS[role]}", "General")
        _calc(ws_db, r, 3, f"={FT_fte[role]}", "#,##0.00")
        _calc(ws_db, r, 4, f"={FT_fte[role]}*{IN['monthly']}", "#,##0")
        _calc(ws_db, r, 5, f"={RT_rate[role]}", "#,##0")
        _calc(ws_db, r, 6, f"={CO_cost[role]}", "#,##0")
        r += 1
    _lbl(ws_db, r, 1, "Total", bold=True)
    _calc(ws_db, r, 6, f"={CO_res}", "#,##0"); r += 2

    _title(ws_db, r, 1, "Effort Breakdown"); r += 1
    _hdr(ws_db, r, 1, "Source"); _hdr(ws_db, r, 2, "Hours"); r += 1
    for label, ref in [
        ("Monitoring Alerts", WL_cat_hrs["alerts"]), ("Service Requests", WL_cat_hrs["service_requests"]),
        ("Incidents", WL_cat_hrs["incidents"]), ("Change Requests", WL_cat_hrs["changes"]),
        ("Patching", PATCH_HRS), ("Additional Activities", AC_total), ("Contingency", EF["contHrs"]),
    ]:
        _lbl(ws_db, r, 1, label); _calc(ws_db, r, 2, f"={ref}", "#,##0.0"); r += 1
    _lbl(ws_db, r, 1, "Total Effort", bold=True); _calc(ws_db, r, 2, f"={EF['total']}", "#,##0.0"); r += 2

    _title(ws_db, r, 1, "Financial Summary"); r += 1
    _hdr(ws_db, r, 1, "Component"); _hdr(ws_db, r, 2, "INR"); r += 1
    for label, ref in [
        ("Total Resource Cost", CO["res"]), ("Additional Expenses", CO["exp"]),
        ("SLA Provision", CO["sla"]), ("TOTAL DELIVERY COST", CO["deliver"]),
        ("Gross Profit", CO["profit"]), ("MONTHLY SELLING PRICE", CO["sell"]),
        ("MONTHLY PRICE incl. Transition", CO["monthly_final"]),
    ]:
        _lbl(ws_db, r, 1, label); _calc(ws_db, r, 2, f"={ref}", "#,##0"); r += 1
    r += 1

    _title(ws_db, r, 1, "Transition & Onboarding Summary"); r += 1
    _hdr(ws_db, r, 1, "Item"); _hdr(ws_db, r, 2, "INR"); r += 1
    for label, ref in [
        ("Total Transition Cost", TR_total), ("Monthly Recurring", TR_monthly),
        ("One-time Fee", TR_onetime), ("Absorbed (discount)", TR_absorb), ("Net Charged", TR_net),
    ]:
        _lbl(ws_db, r, 1, label); _calc(ws_db, r, 2, f"={ref}", "#,##0"); r += 1
    for nm, ref in TR_phase_refs.items():
        _lbl(ws_db, r, 1, f"  · {nm}"); _calc(ws_db, r, 2, f"={ref}", "#,##0"); r += 1

    # ════════════════════════════════════════════════════════════════════════
    # SUMMARY / COVER  (client-facing) — built last, placed first
    # ════════════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet(S_SUM, 0)
    for col, w in (("A", 38), ("B", 20), ("C", 20)):
        ws_sum.column_dimensions[col].width = w
    try:
        _logo = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                             "assets", "nagarro_logo.png")
        if os.path.exists(_logo):
            from openpyxl.drawing.image import Image as XLImage
            _img = XLImage(_logo); _img.width = 140; _img.height = 44
            ws_sum.add_image(_img, "C1")
    except Exception:
        pass
    ws_sum["A1"] = APP_NAME; ws_sum["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws_sum["A2"] = "Managed Services — Cost & Price Summary"
    ws_sum["A2"].font = Font(italic=True, size=10, color=GREY)
    ws_sum["A3"] = f"Generated {date.today()}"; ws_sum["A3"].font = Font(size=9, color=GREY)
    repcur = str(ss.get("reporting_currency") or "INR")
    _treat_label = {"recurring": "Recover via monthly recurring charges",
                    "one_time": "Recover as a one-time fee",
                    "absorb": "Absorbed (discount — net ₹0)"}
    treat_txt = _treat_label.get(tp.get("treatment"), "—") if tr_enabled else "Not included"

    r = 5
    _title(ws_sum, r, 1, "Engagement"); r += 1
    for label, ref in [("Customer / RFP", IN["project"]), ("Prepared by", IN["prep"]),
                       ("Delivery country", IN["country"]), ("Delivery location", IN["location"]),
                       ("Coverage model", IN["covModel"]), ("Reporting currency", IN["repcur"])]:
        _lbl(ws_sum, r, 1, label); _calc(ws_sum, r, 2, f"={ref}", "General"); r += 1
    _lbl(ws_sum, r, 1, "Transition treatment"); ws_sum.cell(r, 2, treat_txt); r += 2

    _title(ws_sum, r, 1, "Commercial Headline"); r += 1
    _hdr(ws_sum, r, 1, "Metric"); _hdr(ws_sum, r, 2, "INR"); _hdr(ws_sum, r, 3, repcur); r += 1

    def head(label, ref, pct=False):
        nonlocal r
        _lbl(ws_sum, r, 1, label)
        _calc(ws_sum, r, 2, f"={ref}", '0.0"%"' if pct else "#,##0")
        if not pct:
            _calc(ws_sum, r, 3, f"=B{r}/IFERROR(VLOOKUP({IN['repcur']},{FX_RANGE},2,FALSE),1)", "#,##0")
        r += 1

    head("Total Delivery Cost", CO["deliver"])
    head("Monthly Selling Price", CO["sell"])
    head("MONTHLY PRICE incl. Transition", CO["monthly_final"])
    head("Transition — total cost", TR_total)
    head("Transition — monthly recurring", TR_monthly)
    head("Transition — net charged", TR_net)
    head("Gross Margin % (on delivery)", IN["margin"], pct=True)
    _lbl(ws_sum, r, 1, "Blended margin % (profit ÷ monthly incl. transition)")
    _calc(ws_sum, r, 2, f"=IF({CO['monthly_final']}>0,{CO['profit']}/{CO['monthly_final']}*100,0)", '0.0"%"'); r += 2

    _title(ws_sum, r, 1, "Key Assumptions"); r += 1
    _hdr(ws_sum, r, 1, "Assumption"); _hdr(ws_sum, r, 2, "Value"); r += 1
    for label, ref, fmt in [
        ("Monthly working hrs / FTE", IN["monthly"], "#,##0"),
        ("Productive utilisation %", IN["util"], '0"%"'),
        ("Contingency %", IN["cont"], '0"%"'),
        ("Coverage multiplier", IN["cov"], "#,##0.00"),
        ("Transition week = hours", IN["trWH"], "#,##0"),
        ("Transition amortisation (months)", IN["trMonths"], "#,##0"),
    ]:
        _lbl(ws_sum, r, 1, label); _calc(ws_sum, r, 2, f"={ref}", fmt); r += 1
    ws_sum.cell(r, 1, "Utilisation legend: 1 = 100%, 0.5 = 50%, 0.25 = 25%, 0 = not used. "
                      "Transition is applied to the price after margin (pass-through). "
                      "Only the Inputs sheet is editable; every other sheet is live formulas.").font = \
        Font(italic=True, size=9, color=GREY)

    # ── Lock everything except the yellow Inputs cells ────────────────────────
    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.selectLockedCells = False
        ws.protection.formatCells = False

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
