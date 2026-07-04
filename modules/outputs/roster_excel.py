"""Excel export for the Shift Plan / Roster Designer.

A presentation artifact (not a recalc model): the roster is a deterministic projection of
the estimate, so we render themed VALUES — reconciliation, the shift grid, and advisories —
ready to drop into a proposal appendix. App theme colours; no logo (parity with the other
formula-driven exports)."""
from __future__ import annotations

import io
from datetime import date
from typing import Any, Dict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import hx

NAVY = hx("navy"); TEAL = hx("teal_dark"); TINT = hx("tint"); MUTED = hx("text_muted")
AMBER = "FBEED9"
_thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _title(ws, r, text, size=13):
    c = ws.cell(r, 1, text); c.font = Font(bold=True, color=NAVY, size=size)


def _kv(ws, r, k, v):
    ws.cell(r, 1, k).font = Font(bold=True, size=10, color=MUTED)
    ws.cell(r, 2, v).font = Font(size=10)


def _hdr_row(ws, r, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(r, j, h); c.fill = _fill(NAVY); c.border = BORDER
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cell(ws, r, j, val, bold=False, right=False, fill=None):
    c = ws.cell(r, j, val); c.border = BORDER
    c.font = Font(size=10, bold=bold)
    c.alignment = Alignment(horizontal="right" if right else "left", vertical="center")
    if fill:
        c.fill = _fill(fill)


def build_roster_workbook(plan: Dict[str, Any], project: str = "") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift Plan"
    ws.sheet_view.showGridLines = False

    r = 1
    _title(ws, r, f"Coverage & Shift Plan{(' — ' + project) if project else ''}"); r += 2
    _kv(ws, r, "Strategy", plan.get("strategy", "Balanced")); r += 1
    _kv(ws, r, "Customer time zone", plan.get("customer_tz", "")); r += 1
    _kv(ws, r, "Delivery time zone", plan.get("delivery_tz", "")); r += 1
    _kv(ws, r, "Business hours", plan.get("business_hours", "")); r += 1
    _kv(ws, r, "Basis", f"{plan.get('fte_basis', 'rounded')} FTE (delivered team)"); r += 1
    _kv(ws, r, "Generated", date.today().isoformat()); r += 2

    # Reconciliation — billed FTE → deployable seats
    _title(ws, r, "FTE → Deployable Seats (reconciliation)", 11); r += 1
    _hdr_row(ws, r, ["Skill", "Level", "Coverage", "Billed FTE", "Seats", "Delta"]); r += 1
    for row in plan.get("reconciliation", []):
        _cell(ws, r, 1, row["skill"])
        _cell(ws, r, 2, row["level"])
        _cell(ws, r, 3, row["coverage"])
        _cell(ws, r, 4, row["fte"], right=True)
        _cell(ws, r, 5, row["seats"], right=True)
        _cell(ws, r, 6, f"+{row['delta']:.2f}" if row["delta"] >= 0 else f"{row['delta']:.2f}", right=True)
        r += 1
    tot = plan.get("totals", {})
    _cell(ws, r, 1, "Total", bold=True, fill=TINT)
    for j in (2, 3):
        _cell(ws, r, j, "", fill=TINT)
    _cell(ws, r, 4, tot.get("delivered_fte", 0), bold=True, right=True, fill=TINT)
    _cell(ws, r, 5, tot.get("deployable_seats", 0), bold=True, right=True, fill=TINT)
    _cell(ws, r, 6, f"+{tot.get('delta', 0):.2f}", bold=True, right=True, fill=TINT)
    r += 2

    # Shift-timing legend
    _title(ws, r, "Shift Timings", 11); r += 1
    _hdr_row(ws, r, ["Shift", f"Customer ({plan.get('customer_tz','')})",
                     f"Delivery ({plan.get('delivery_tz','')})"]); r += 1
    for t in plan.get("shift_timings", []):
        _cell(ws, r, 1, t["label"])
        _cell(ws, r, 2, t["customer"], right=True)
        _cell(ws, r, 3, t["delivery"], right=True)
        r += 1
    r += 1

    # Weekly roster (person × weekday) — the requested sample format:
    # Employee | Role | Filter | Mon…Sun (anonymous seats, no rate column).
    _title(ws, r, "Weekly Roster", 11); r += 1
    days = plan.get("days", ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
    header_row = r
    _hdr_row(ws, r, ["Employee", "Role", "Filter"] + days); r += 1
    cell_bg = {"Morning": "D6F0ED", "Evening": "A8DDD8", "Night": "1A5F6A",
               "Day": "EAF3F4", "On-Call": AMBER}
    for p in plan.get("people", []):
        _cell(ws, r, 1, p["employee"])
        _cell(ws, r, 2, p["role"])
        _cell(ws, r, 3, "")   # Filter helper column (Excel autofilter)
        for j, v in enumerate(p["cells"], start=4):
            fill = cell_bg.get(v)
            c = ws.cell(r, j, v or ""); c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9, color=("FFFFFF" if v == "Night" else "1B2A3A"))
            if fill:
                c.fill = _fill(fill)
        r += 1
    # Autofilter across the roster header (the "Filter" column mirrors the sample).
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(3 + len(days))}{r - 1}"
    r += 1

    # Roster notes
    for n in plan.get("roster_notes", []):
        c = ws.cell(r, 1, "• " + n); c.font = Font(size=9, italic=True, color=MUTED)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3 + len(days))
        r += 1
    r += 1

    # Advisories
    advisories = plan.get("advisories", [])
    if advisories:
        _title(ws, r, "Coverage Advisories (informational — do not affect commercials)", 11); r += 1
        for a in advisories:
            c = ws.cell(r, 1, "⚠  " + a); c.fill = _fill(AMBER)
            c.font = Font(size=10); c.alignment = Alignment(wrap_text=True, vertical="center")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3 + len(days))
            r += 1

    widths = [20, 26, 8] + [11] * len(days)
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
