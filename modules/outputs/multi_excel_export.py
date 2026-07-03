"""
Multi-skill Excel export — a multi-sheet workbook built from the multi-skill engine.

State-driven (no session reads), so it's unit-testable: pass a `state` dict (the same
shape `_build_multi_state()` produces) or let it read the live one. The numbers are
written straight from `compute_multi_skill_model`, so the workbook equals the engine
(a faithful values report, not a live-formula spreadsheet).

Sheets: Executive Summary · Skills · Effort Build-up · Team (FTE) · Rates ·
Optimization · Workload Detail · Inputs.
Styling helpers are reused from the single-mode exporter so the look matches.
"""
import io
from datetime import date

import openpyxl

from config.settings import APP_NAME, hx
from modules.calculations.engine import compute_multi_skill_model, calc_patching_effort
from modules.outputs.excel_export import (
    _hrow, _drow, _title, _add_logo, NAVY, BLUE, LB, ACCENT,
    F_INR, F_NUM1, F_FTE, F_RAW, F_PCT,
)

LEVELS4 = ["L1", "L2", "L3", "Architect"]


def _sum_role_hours(ps):
    return sum(ps["role_hours"].get(l, 0.0) for l in LEVELS4)


# ── Sheets ───────────────────────────────────────────────────────────────────
def _exec(wb, model, baseline, state):
    ws = wb.create_sheet("Executive Summary")
    _add_logo(ws)
    _title(ws, f"{APP_NAME} — Multi-skill Executive Summary")
    cr, pr = model["cost_result"], model["price_result"]
    staffed = sum(_sum_role_hours(ps) for ps in model["per_skill"].values())
    _hrow(ws, 5, ["Metric", "Value", "Unit"], [40, 22, 14])
    rows = [
        ("Skills in scope", len(model["per_skill"]), ""),
        ("Delivery location",
         (state.get("delivery_country") or "—") +
         (f" / {state.get('delivery_location')}" if state.get("delivery_location") else ""), ""),
        ("TOTAL STAFFED EFFORT (L1–Architect)", round(staffed, 1), "Hrs/Month"),
        ("SDM hours", round(model["sdm_hours"], 1), "Hrs/Month"),
        ("TOTAL FTE", round(model["total_fte"], 1), "FTE"),
        ("Resource cost", cr.get("resource_cost", 0), "INR"),
        ("Delivery cost", cr.get("total_delivery_cost", 0), "INR"),
        ("Gross margin", pr.get("margin_pct", 0), "%"),
        ("MONTHLY SELLING PRICE", pr.get("selling_price", 0), "INR"),
        ("Gross profit", pr.get("gross_profit", 0), "INR"),
    ]
    saved = baseline["total_fte"] - model["total_fte"]
    if saved > 1e-9:
        rows += [
            ("— Optimisation applied —", "", ""),
            ("Baseline team (no sharing)", round(baseline["total_fte"], 1), "FTE"),
            ("FTE saved by sharing", round(saved, 1), "FTE"),
            ("Cost saved", baseline["total_resource_cost"] - model["total_resource_cost"], "INR"),
        ]
    fmt = {"": None, "Hrs/Month": F_NUM1, "FTE": F_FTE, "INR": F_INR, "%": F_PCT}
    for i, (m, v, u) in enumerate(rows, 6):
        _drow(ws, i, [m, v, u], total=str(m).isupper(), fmts=[None, fmt.get(u), None])


def _skills(wb, model, names):
    ws = wb.create_sheet("Skills")
    _title(ws, "Per-skill Effort, FTE & Cost")
    _hrow(ws, 5, ["Skill", "Family", "Coverage", "L1 hrs", "L2 hrs", "L3 hrs", "Arch hrs",
                  "Staffed hrs", "Monthly Cost"],
          [22, 12, 10, 10, 10, 10, 10, 12, 16])
    fmts = [None, None, None, F_NUM1, F_NUM1, F_NUM1, F_NUM1, F_NUM1, F_INR]
    r = 6
    tot = {l: 0.0 for l in LEVELS4}
    tcost = tstaff = 0.0
    for sid, ps in model["per_skill"].items():
        rh = ps["role_hours"]
        staffed = _sum_role_hours(ps)
        for l in LEVELS4:
            tot[l] += rh.get(l, 0.0)
        tstaff += staffed
        tcost += ps.get("cost", 0.0)
        _drow(ws, r, [names.get(sid, sid), ps["genus_category"], ps["coverage_model"],
                      round(rh["L1"], 1), round(rh["L2"], 1), round(rh["L3"], 1),
                      round(rh["Architect"], 1), round(staffed, 1), round(ps.get("cost", 0.0))], fmts=fmts)
        r += 1
    _drow(ws, r, ["TOTAL", "", ""] + [round(tot[l], 1) for l in LEVELS4] +
          [round(tstaff, 1), round(tcost)], total=True, fmts=fmts)


def _buildup(wb, model, names):
    ws = wb.create_sheet("Effort Build-up")
    _title(ws, "Raw → Buffered → Final (per skill × level)")
    _hrow(ws, 5, ["Skill", "Level", "Raw hrs", "Buffer %", "Buffered hrs", "Final hrs",
                  "Raw FTE", "Final FTE"], [22, 10, 11, 9, 12, 11, 10, 10])
    fmts = [None, None, F_NUM1, F_PCT, F_NUM1, F_NUM1, F_RAW, F_FTE]
    r = 6
    for sid, ps in model["per_skill"].items():
        for lvl in LEVELS4:
            d = ps["breakdown"][lvl]
            if d["raw"] <= 1e-9 and d["final"] <= 1e-9:
                continue
            _drow(ws, r, [names.get(sid, sid), lvl, round(d["raw"], 1), round(d["buffer_pct"], 0),
                          round(d["buffered"], 1), round(d["final"], 1),
                          round(d["fte_raw"], 3), round(d["fte_staffed"], 1)], fmts=fmts)
            r += 1


def _team(wb, model, names):
    ws = wb.create_sheet("Team (FTE)")
    _title(ws, "Team composition — Raw vs Final FTE by skill × level")
    sdm = next((x for x in model["resources"] if x["level"] == "SDM"), None)
    sdm_raw = float(sdm["raw_fte"]) if sdm else 0.0
    sdm_fin = float(sdm["fte"]) if sdm else 0.0

    def block(row0, title, kind, sdm_val):
        ws.cell(row0, 1, title).font = openpyxl.styles.Font(name="Calibri", bold=True, color=NAVY, size=11)
        _hrow(ws, row0 + 1, ["Skill", "L1", "L2", "L3", "Architect", "Total"],
              [22, 10, 10, 10, 12, 10])
        fmts = [None, F_FTE, F_FTE, F_FTE, F_FTE, F_FTE]
        r = row0 + 2
        col_tot = {l: 0.0 for l in LEVELS4}
        grand = 0.0
        for sid, ps in model["per_skill"].items():
            vals = []
            rt = 0.0
            for l in LEVELS4:
                v = ps["breakdown"][l]["fte_raw"] if kind == "raw" else ps["fte_by_level"][l]
                col_tot[l] += v
                rt += v
                vals.append(round(v, 2))
            grand += rt
            _drow(ws, r, [names.get(sid, sid)] + vals + [round(rt, 2)], fmts=fmts)
            r += 1
        if sdm_val > 1e-9:
            _drow(ws, r, ["SDM (engagement)", "", "", "", "", round(sdm_val, 2)], fmts=fmts)
            r += 1
        _drow(ws, r, ["GRAND TOTAL"] + [round(col_tot[l], 2) for l in LEVELS4] +
              [round(grand + sdm_val, 2)], total=True, fmts=fmts)
        return r + 2

    nxt = block(5, "Raw FTE (exact, pre-pooling)", "raw", sdm_raw)
    block(nxt, "Final FTE (delivered, pooled-aware)", "final", sdm_fin)


def _rates(wb, state):
    ws = wb.create_sheet("Rates")
    _title(ws, "Resolved hourly rates (INR)")
    rbc = state.get("rates_by_category", {}) or {}
    _hrow(ws, 5, ["Family", "L1 /hr", "L2 /hr", "L3 /hr", "Architect /hr"], [16, 12, 12, 12, 14])
    fmts = [None, F_INR, F_INR, F_INR, F_INR]
    r = 6
    for fam in ("InfraOps", "CloudOps"):
        band = rbc.get(fam, {}) or {}
        _drow(ws, r, [fam] + [round(float(band.get(l, 0) or 0)) for l in LEVELS4], fmts=fmts)
        r += 1
    _drow(ws, r, ["SDM (engagement)", round(float(state.get("sdm_rate_inr", 0) or 0)), "", "", ""],
          fmts=fmts)


def _optimization(wb, baseline, current, opt, names):
    ws = wb.create_sheet("Optimization")
    _title(ws, "AI Team Optimizer — savings & recommended moves")
    _hrow(ws, 5, ["Metric", "Baseline (no sharing)", "Applied", "Saving"], [26, 20, 18, 16])
    fmts = [None, F_FTE, F_FTE, F_FTE]
    _drow(ws, 6, ["Total FTE", round(baseline["total_fte"], 1), round(current["total_fte"], 1),
                  round(baseline["total_fte"] - current["total_fte"], 1)], fmts=fmts)
    cfmts = [None, F_INR, F_INR, F_INR]
    _drow(ws, 7, ["Resource cost / mo", round(baseline["total_resource_cost"]),
                  round(current["total_resource_cost"]),
                  round(baseline["total_resource_cost"] - current["total_resource_cost"])], fmts=cfmts)
    _drow(ws, 8, ["Selling price / mo", round(baseline["price_result"]["selling_price"]),
                  round(current["price_result"]["selling_price"]),
                  round(baseline["price_result"]["selling_price"] - current["price_result"]["selling_price"])],
          fmts=cfmts)

    _title_row = 10
    ws.cell(_title_row, 1, "Recommended moves (advisory)").font = openpyxl.styles.Font(
        name="Calibri", bold=True, color=NAVY, size=11)
    _hrow(ws, _title_row + 1, ["Skills", "Level", "Coverage", "FTE saved", "Cost saved",
                               "Cross-family", "Key-person risk"], [30, 10, 10, 11, 14, 12, 14])
    mfmts = [None, None, None, F_FTE, F_INR, None, None]
    r = _title_row + 2
    for s in opt.get("suggestions", []):
        _drow(ws, r, [" + ".join(s["skill_names"]), s["level"], s["coverage_model"],
                      round(s["fte_saved"], 1), round(s.get("cost_saved", 0)),
                      "Yes" if s.get("cross_family") else "", "Yes" if s["key_person_risk"] else ""],
              fmts=mfmts)
        r += 1
    if not opt.get("suggestions"):
        ws.cell(r, 1, "No cross-skill sharing opportunities found for the current setup.")


def _workload(wb, state, names):
    ws = wb.create_sheet("Workload Detail")
    _title(ws, "Inputs — Tickets, Patching & Additional Activities")
    skills = state.get("skills", [])
    row = 5
    # Tickets
    ws.cell(row, 1, "Tickets").font = openpyxl.styles.Font(name="Calibri", bold=True, color=NAVY, size=11)
    _hrow(ws, row + 1, ["Skill", "Category", "Classification", "Count", "Min/Ticket", "L1 %", "L2 %", "L3 %"],
          [22, 18, 14, 10, 11, 8, 8, 8])
    tf = [None, None, None, F_INR, F_NUM1, F_PCT, F_PCT, F_PCT]
    r = row + 2
    cats = [("alerts", "Monitoring Alerts"), ("service_requests", "Service Requests"),
            ("incidents", "Incidents"), ("changes", "Change Requests")]
    for sk in skills:
        wl = sk.get("workload", {}) or {}
        for ck, cl in cats:
            for cls, rw in (wl.get(ck, {}) or {}).items():
                if not rw or (rw.get("count", 0) or 0) <= 0:
                    continue
                _drow(ws, r, [sk.get("name"), cl, cls, int(rw.get("count", 0) or 0),
                              float(rw.get("minutes", 0) or 0), float(rw.get("L1_pct", 0) or 0),
                              float(rw.get("L2_pct", 0) or 0), float(rw.get("L3_pct", 0) or 0)], fmts=tf)
                r += 1
    # Patching
    r += 1
    ws.cell(r, 1, "Patching").font = openpyxl.styles.Font(name="Calibri", bold=True, color=NAVY, size=11)
    _hrow(ws, r + 1, ["Skill", "Servers", "Method", "Handled by", "Hrs/Month"], [22, 10, 12, 12, 12])
    pf = [None, F_INR, None, None, F_NUM1]
    r += 2
    for sk in skills:
        p = sk.get("patching") or {}
        if not p.get("included"):
            continue
        res = calc_patching_effort(True, p.get("num_servers", 0) or 0, p.get("method") or "Manual",
                                   p.get("manual_effort_per_server", 45) or 45,
                                   p.get("auto_effort_per_server", 30) or 30,
                                   error_rate_pct=p.get("error_rate_pct", 0) or 0)
        _drow(ws, r, [sk.get("name"), int(p.get("num_servers", 0) or 0), p.get("method"),
                      p.get("patching_role"), round(res["hours"], 1)], fmts=pf)
        r += 1
    # Activities
    r += 1
    ws.cell(r, 1, "Additional Activities").font = openpyxl.styles.Font(name="Calibri", bold=True, color=NAVY, size=11)
    _hrow(ws, r + 1, ["Skill", "Activity", "Auto", "Hrs/Month", "L1 %", "L2 %", "L3 %", "Arch %"],
          [22, 24, 7, 11, 8, 8, 8, 8])
    af = [None, None, None, F_NUM1, F_PCT, F_PCT, F_PCT, F_PCT]
    r += 2
    for sk in skills:
        for a in (sk.get("activities") or []):
            d = a.get("dist", {}) or {}
            _drow(ws, r, [sk.get("name"), a.get("name"), "Yes" if a.get("auto") else "",
                          float(a.get("hours", 0) or 0)] +
                  [float(d.get(l, 0) or 0) for l in LEVELS4], fmts=af)
            r += 1


def _inputs(wb, state):
    ws = wb.create_sheet("Inputs")
    _title(ws, "Engagement inputs")
    _hrow(ws, 5, ["Input", "Value", "Unit"], [34, 20, 12])
    rows = [
        ("Monthly working hours / FTE", state.get("monthly_working_hours"), "Hrs"),
        ("Productive utilisation", state.get("productive_utilisation"), "%"),
        ("Contingency", state.get("contingency_pct"), "%"),
        ("SDM overhead", state.get("sdm_overhead_pct"), "%"),
        ("Target margin", state.get("target_margin_pct"), "%"),
        ("Context-switch penalty", state.get("context_switch_pct"), "%"),
        ("Enforce 24×7 shift minimums", "Yes" if state.get("enforce_min_shift") else "No", ""),
        ("FTE basis", state.get("fte_basis"), ""),
        ("Delivery country", state.get("delivery_country"), ""),
        ("Delivery location", state.get("delivery_location") or "—", ""),
    ]
    fmt = {"": None, "Hrs": F_NUM1, "%": F_PCT}
    for i, (k, v, u) in enumerate(rows, 6):
        _drow(ws, i, [k, v, u], fmts=[None, fmt.get(u), None])


def _raw_vs_rounded(wb, state):
    ws = wb.create_sheet("Raw vs Rounded")
    _title(ws, "Raw vs Rounded — two estimate versions")
    raw = compute_multi_skill_model({**state, "fte_basis": "raw"})
    rnd = compute_multi_skill_model({**state, "fte_basis": "rounded"})
    chosen = "Raw" if state.get("fte_basis") == "raw" else "Rounded"
    ws.cell(4, 1, f"Reported / priced basis: {chosen}.  Raw = exact fractional demand (theoretical "
                  "minimum; assumes perfect pooling).  Rounded = delivered team (each skill × level "
                  "rounded up to 0.5, minimum 0.5).")
    _hrow(ws, 6, ["Metric", "Raw (theoretical)", "Rounded (delivered)", "Δ (Rounded − Raw)"],
          [30, 20, 20, 20])
    cr_r, cr_f, pr_r, pr_f = raw["cost_result"], rnd["cost_result"], raw["price_result"], rnd["price_result"]
    rows = [
        ("Total FTE", raw["total_fte"], rnd["total_fte"], F_RAW),
        ("Resource cost / mo", raw["total_resource_cost"], rnd["total_resource_cost"], F_INR),
        ("Delivery cost / mo", cr_r["total_delivery_cost"], cr_f["total_delivery_cost"], F_INR),
        ("Selling price / mo", pr_r["selling_price"], pr_f["selling_price"], F_INR),
        ("Gross profit / mo", pr_r["gross_profit"], pr_f["gross_profit"], F_INR),
    ]
    for i, (m, rv, fv, f) in enumerate(rows, 7):
        _drow(ws, i, [m, rv, fv, fv - rv], fmts=[None, f, f, f])


def _live_model(wb, model, state):
    """Formula-driven, editable multi-skill model (pragmatic Phase B). Teal cells are
    editable inputs; every result is a live Excel formula that reproduces the engine.
    Per skill × category: Volume, blended AHT and blended L1/L2/L3 split (linear in volume,
    so blended reproduces the engine exactly); + per-level buffer, non-ticket ('Other') hrs,
    Architect %, rates; + engagement monthly hrs / utilisation / contingency / SDM% / margin.
    Pooling/optimisation is NOT live here (see the Optimization sheet) — this is the un-pooled
    build; a note flags it when sharing is applied."""
    from openpyxl.styles import Font, Protection
    from openpyxl.utils import get_column_letter
    from modules.outputs.excel_model import _edit, _calc, _hdr, _lbl, _fill, _aref, YEL, BORDER, NAVY
    from modules.calculations.engine import calc_coverage_multiplier

    ws = wb.create_sheet("Live Model")
    for col, w in (("A", 30), ("B", 12), ("C", 12), ("D", 12), ("E", 13), ("F", 12), ("G", 15)):
        ws.column_dimensions[col].width = w
    ws.cell(1, 1, "Live Model — editable & formula-driven").font = Font(bold=True, color=NAVY, size=13)
    lg = ws.cell(2, 1, "  Editable input  "); lg.fill = _fill(YEL); lg.font = Font(italic=True, bold=True, size=9); lg.border = BORDER
    ws.cell(2, 2, "Teal cells are editable — change them and the workbook recalculates. Everything "
                  "else is a locked live formula. Cost uses Rounded FTE (delivered).").font = Font(italic=True, size=9, color="6B7B7B")

    cats = [("alerts", "Monitoring Alerts"), ("service_requests", "Service Requests"),
            ("incidents", "Incidents"), ("changes", "Change Requests")]
    LV = ["L1", "L2", "L3"]
    raw_basis = str(state.get("fte_basis", "rounded")).lower() == "raw"

    def _fte_formula(fin_ref, covx):
        raw = f"{fin_ref}/{PRD}{covx}"
        return f"={raw}" if raw_basis else f"=IF({fin_ref}>0,MAX(CEILING({raw},0.5),0.5),0)"

    # ── Engagement inputs ──
    r = 4
    _hdr(ws, r, 1, "Engagement inputs"); r += 1
    monthly = float(state.get("monthly_working_hours", 160.0) or 160.0)
    util = float(state.get("productive_utilisation", 75.0) or 75.0)
    cont = float(state.get("contingency_pct", 10.0) or 0.0)
    sdmp = float(state.get("sdm_overhead_pct", 0.0) or 0.0)
    margin = float(state.get("target_margin_pct", 0.0) or 0.0)
    _lbl(ws, r, 1, "Monthly working hrs / FTE"); MON = _edit(ws, r, 2, monthly, "#,##0"); r += 1
    _lbl(ws, r, 1, "Productive utilisation %"); UTL = _edit(ws, r, 2, util, "#,##0"); r += 1
    _lbl(ws, r, 1, "Productive hrs / FTE"); PRD = _calc(ws, r, 2, f"={MON}*{UTL}/100", "#,##0.0"); r += 1
    _lbl(ws, r, 1, "Contingency %"); CON = _edit(ws, r, 2, cont, "#,##0"); r += 1
    _lbl(ws, r, 1, "SDM allocation (% of one SDM)"); SDM = _edit(ws, r, 2, sdmp, "#,##0"); r += 1
    _lbl(ws, r, 1, "Target margin %"); MRG = _edit(ws, r, 2, margin, "#,##0"); r += 2

    skill_cost_refs, skill_base_refs = [], []
    for sk in state.get("skills", []):
        ps = model["per_skill"].get(sk["id"], {})
        fam = sk.get("genus_category", "InfraOps")
        rates = (state.get("rates_by_category", {}) or {}).get(fam, {}) or {}
        cov = calc_coverage_multiplier(sk.get("coverage_model") or "8×5",
                                       state.get("custom_hours_per_day", 8), state.get("custom_days_per_week", 5))
        rb = sk.get("role_buffers") or {}
        wl = sk.get("workload", {}) or {}

        _hdr(ws, r, 1, f"{sk.get('name') or sk['id']}  ·  {fam}  ·  {sk.get('coverage_model','')}")
        for cc in range(2, 7):
            ws.cell(r, cc).fill = _fill(NAVY)
        r += 1
        _lbl(ws, r, 1, "Coverage multiplier"); COV = _edit(ws, r, 2, round(cov, 3), "#,##0.000")
        _lbl(ws, r, 4, "Architect %"); ARP = _edit(ws, r, 5, float(sk.get("architect_pct", 0) or 0) if sk.get("has_architect") else 0.0, "#,##0"); r += 1
        # Category rows → refs for volume / blended AHT / blended split
        ws.cell(r, 1, "Category").font = Font(bold=True, size=9)
        for cc, t in ((2, "Volume"), (3, "AHT min"), (4, "L1 %"), (5, "L2 %"), (6, "L3 %")):
            ws.cell(r, cc, t).font = Font(bold=True, size=9)
        r += 1
        cat_ref = {}          # cat -> (vol, aht, {L: splitref})
        for ckey, clabel in cats:
            rows = wl.get(ckey, {}) or {}
            total_cnt = sum(float((rw or {}).get("count", 0) or 0) for rw in rows.values())
            tot_min = sum(float((rw or {}).get("count", 0) or 0) * float((rw or {}).get("minutes", 0) or 0) for rw in rows.values())
            blend_aht = (tot_min / total_cnt) if total_cnt > 0 else 0.0
            split = {}
            for L in LV:
                num = sum(float((rw or {}).get("count", 0) or 0) * float((rw or {}).get("minutes", 0) or 0)
                          * float((rw or {}).get(f"{L}_pct", 0) or 0) / 100.0 for rw in rows.values())
                split[L] = (num / tot_min * 100.0) if tot_min > 0 else 0.0
            _lbl(ws, r, 1, clabel)
            v = _edit(ws, r, 2, round(total_cnt), "#,##0")
            a = _edit(ws, r, 3, round(blend_aht, 1), "#,##0.0")
            sp = {L: _edit(ws, r, 3 + i, round(split[L], 1), "#,##0.0") for i, L in enumerate(LV, start=1)}
            cat_ref[ckey] = (v, a, sp)
            r += 1
        # Per-level rows
        ws.cell(r, 1, "").value = None
        hdr_r = r
        _lbl(ws, r, 1, "By level");
        for i, L in enumerate(["L1", "L2", "L3", "Architect"], start=2):
            ws.cell(r, i, L).font = Font(bold=True, size=9)
        r += 1
        # Ticket raw hours per level (formula from category rows)
        _lbl(ws, r, 1, "Ticket hours (raw)")
        tk = {}
        for i, L in enumerate(LV, start=2):
            terms = "+".join(f"{v}*{a}/60*{sp[L]}/100" for (v, a, sp) in cat_ref.values())
            tk[L] = _calc(ws, r, i, f"={terms}", "#,##0.0")
        ws.cell(r, 5, "—")
        r += 1
        # Buffer % (editable)
        _lbl(ws, r, 1, "Buffer %"); buf = {}
        for i, L in enumerate(["L1", "L2", "L3", "Architect"], start=2):
            buf[L] = _edit(ws, r, i, float(rb.get(L, 20) or 0), "#,##0")
        r += 1
        # Other (non-ticket) hrs per level = engine breakdown raw - ticket raw (patch+activities+arch)
        _lbl(ws, r, 1, "Other (non-ticket) hrs")
        bd = ps.get("breakdown", {})
        base_effort = float(ps.get("total_effort", 0) or 0) / (1 + cont / 100.0)
        oth = {}
        for i, L in enumerate(["L1", "L2", "L3"], start=2):
            ticket_raw_L = sum(float((rw or {}).get("count", 0) or 0) * float((rw or {}).get("minutes", 0) or 0) / 60.0
                               * float((rw or {}).get(f"{L}_pct", 0) or 0) / 100.0
                               for ck in [c for c, _ in cats] for rw in (wl.get(ck, {}) or {}).values())
            extra = float(bd.get(L, {}).get("raw", 0) or 0) - ticket_raw_L
            oth[L] = _edit(ws, r, i, round(max(extra, 0.0), 1), "#,##0.0")
        arch_raw_seed = base_effort * (float(sk.get("architect_pct", 0) or 0) / 100.0 if sk.get("has_architect") else 0.0)
        extra_arch = float(bd.get("Architect", {}).get("raw", 0) or 0) - arch_raw_seed
        oth["Architect"] = _edit(ws, r, 5, round(max(extra_arch, 0.0), 1), "#,##0.0")
        r += 1
        # base effort formula (tickets + all other)
        base_ref = f"({'+'.join(f'{v}*{a}/60' for (v, a, sp) in cat_ref.values())}+{oth['L1']}+{oth['L2']}+{oth['L3']}+{oth['Architect']})"
        # Final hours per level
        _lbl(ws, r, 1, "Final hours")
        fin = {}
        for i, L in enumerate(LV, start=2):
            fin[L] = _calc(ws, r, i, f"=({tk[L]}*(1+{buf[L]}/100)+{oth[L]})*(1+{CON}/100)", "#,##0.0")
        fin["Architect"] = _calc(ws, r, 5, f"=({base_ref}*{ARP}/100*(1+{buf['Architect']}/100)+{oth['Architect']})*(1+{CON}/100)", "#,##0.0")
        r += 1
        # FTE per level (coverage on L1/L2; basis = the estimate's fte_basis)
        _lbl(ws, r, 1, "FTE (Raw)" if raw_basis else "FTE (Rounded)")
        fte = {}
        for L, i in (("L1", 2), ("L2", 3), ("L3", 4), ("Architect", 5)):
            covx = f"*{COV}" if L in ("L1", "L2") else ""
            fte[L] = _calc(ws, r, i, _fte_formula(fin[L], covx), "#,##0.00")
        r += 1
        # Rate per level (editable) + cost
        _lbl(ws, r, 1, "Rate INR/hr"); rt = {}
        for L, i in (("L1", 2), ("L2", 3), ("L3", 4), ("Architect", 5)):
            rt[L] = _edit(ws, r, i, round(float(rates.get(L, 0) or 0)), "#,##0")
        r += 1
        _lbl(ws, r, 1, "Cost INR/mo")
        cost_terms = []
        for L, i in (("L1", 2), ("L2", 3), ("L3", 4), ("Architect", 5)):
            cref = _calc(ws, r, i, f"={fte[L]}*{MON}*{rt[L]}", "#,##0")
            cost_terms.append(cref)
        sc = _calc(ws, r, 6, f"={'+'.join(cost_terms)}", "#,##0")
        ws.cell(r, 7, round(float(ps.get("cost", 0) or 0))).font = Font(italic=True, color="6B7B7B")  # app value
        ws.cell(hdr_r - 1, 7, "app →").font = Font(italic=True, size=8, color="6B7B7B")
        skill_cost_refs.append(sc)
        skill_base_refs.append(base_ref)
        r += 2

    # ── Engagement totals ──
    _hdr(ws, r, 1, "Engagement totals"); r += 1
    res = _calc(ws, r, 2, f"={'+'.join(skill_cost_refs) if skill_cost_refs else '0'}", "#,##0")
    _lbl(ws, r, 1, "Resource cost (skills)"); r += 1
    # SDM (Option A): a fixed allocation of ONE SDM — FTE = SDM%/100 (not a % of effort, not
    # rounded). Hours = FTE × productive; cost = FTE × monthly × rate.
    _lbl(ws, r, 1, "SDM FTE (= SDM% of one SDM)"); sdm_fte = _calc(ws, r, 2, f"={SDM}/100", "#,##0.00"); r += 1
    _lbl(ws, r, 1, "SDM hours / mo"); _calc(ws, r, 2, f"={sdm_fte}*{PRD}", "#,##0.0"); r += 1
    _lbl(ws, r, 1, "SDM rate INR/hr")
    SDR = _edit(ws, r, 2, round(float(state.get("sdm_rate_inr", 0) or 0)), "#,##0"); r += 1
    _lbl(ws, r, 1, "SDM cost / mo"); sdm_cost = _calc(ws, r, 2, f"={sdm_fte}*{MON}*{SDR}", "#,##0"); r += 1
    _lbl(ws, r, 1, "Total resource cost")
    tot_cost = _calc(ws, r, 2, f"={res}+{sdm_cost}", "#,##0")
    ws.cell(r, 3, round(float(model.get("total_resource_cost", 0) or 0))).font = Font(italic=True, color="6B7B7B")
    ws.cell(r, 4, "← app").font = Font(italic=True, size=8, color="6B7B7B"); r += 1
    _lbl(ws, r, 1, "Selling price / mo")
    price = _calc(ws, r, 2, f"=IF({MRG}<100,{tot_cost}/(1-{MRG}/100),{tot_cost})", "#,##0")
    ws.cell(r, 3, round(float(model["price_result"].get("selling_price", 0) or 0))).font = Font(italic=True, color="6B7B7B")
    ws.cell(r, 4, "← app").font = Font(italic=True, size=8, color="6B7B7B"); r += 2
    if state.get("resource_sharing"):
        ws.cell(r, 1, "Note: resource sharing (pooling) is applied in the app but NOT modelled live here "
                      "— this sheet is the un-pooled build. See the Optimization sheet for the pooled delivered team.").font = \
            Font(italic=True, size=9, color="B8860B")

    # lock everything except the editable (teal, unlocked) cells
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.sheet_view.showGridLines = False


def generate_multi_excel_report(state=None) -> bytes:
    """Build the multi-skill workbook and return .xlsx bytes. `state` defaults to the
    live `build_multi_model_state()`; pass a dict to render without the UI (tests)."""
    if state is None:
        from modules.state.multi_state import build_multi_model_state
        state = build_multi_model_state()
    model = compute_multi_skill_model(state)
    baseline = compute_multi_skill_model({**state, "resource_sharing": []})
    try:
        from modules.optimize.team_optimizer import optimize_team
        opt = optimize_team({**state, "resource_sharing": []})
    except Exception:
        opt = {"suggestions": []}
    names = {s["id"]: (s.get("name") or s["id"]) for s in state.get("skills", [])}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _exec(wb, model, baseline, state)
    _live_model(wb, model, state)
    _skills(wb, model, names)
    _buildup(wb, model, names)
    _team(wb, model, names)
    _rates(wb, state)
    _raw_vs_rounded(wb, state)
    _optimization(wb, baseline, model, opt, names)
    _workload(wb, state, names)
    _inputs(wb, state)

    colors = [NAVY, hx("teal_dark"), BLUE, ACCENT, hx("success"), hx("primary"), hx("text_muted"), NAVY]
    for i, ws in enumerate(wb.worksheets):
        ws.sheet_properties.tabColor = colors[i % len(colors)]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
